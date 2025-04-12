from sqlalchemy import create_engine
import pandas as pd
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from multiprocessing import Pool, cpu_count
from functools import partial

# SQL 操作部分
def get_db_engine(host, port, user, password, database):
    """创建 SQLAlchemy 引擎"""
    try:
        connection_string = f"mysql+pymysql://{user}:{password}@{host}:{port}/{database}"
        return create_engine(connection_string, pool_size=5, pool_recycle=3600)
    except Exception as e:
        print(f"创建数据库引擎出错: {e}")
        return None

def fetch_member_ids(engine, database, table_name):
    """查询满足条件的 member_ids"""
    sql = f"""
    SELECT member_id
    FROM {database}.{table_name}
    WHERE DATE(created_at) >= '2024-12-30'
    GROUP BY member_id
    HAVING SUM(order_amount) > 0
    """
    try:
        return tuple(pd.read_sql(sql, engine)['member_id'])
    except Exception as e:
        print(f"获取 member_ids 出错: {e}")
        return ()

def fetch_main_data(engine, member_ids, database, table_name):
    """执行主查询并返回数据"""
    base_sql = f"""
    SELECT
        DATE(created_at) AS created_date,
        site_id, member_id, member_username, id, created_at,
        SUM(order_amount) AS total_order_amount,
        CASE WHEN pay_type = 1006 THEN '好博-MPay支付-MPay代收' ELSE pay_type END AS '支付方式',
        pay_channel, order_status
    FROM {database}.{table_name}
    WHERE created_at > '2024-12-30 00:00:00'
    AND created_at < '2025-01-01 23:59:59'
    AND site_id = 2000
    AND order_status = -1
    {{}}
    GROUP BY created_date, member_id, member_username, id, created_at, pay_channel, order_status, site_id
    ORDER BY created_at ASC, total_order_amount DESC
    LIMIT 9999999999
    """
    sql = base_sql.format(f"AND member_id IN {member_ids}" if member_ids else "")
    try:
        return pd.read_sql(sql, engine, chunksize=500000)  # 保持低内存占用
    except Exception as e:
        print(f"执行主查询出错: {e}")
        return None

# Excel 操作部分
def generate_excel_filename(database, table_name):
    """生成 Excel 文件名"""
    now = datetime.datetime.now().strftime("%Y-%m-%d_%H.%M")
    return f"{database}_{table_name}_{now}.xlsx"

def write_chunk_to_excel(args):
    """多进程写入 Excel 的单块数据"""
    df_chunk, sheet_name, excel_filename = args
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a' if sheet_name != 'Sheet1' else 'w')
    df_chunk.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()
    return len(df_chunk)

def format_excel(excel_filename, column_count):
    """格式化 Excel 文件"""
    workbook = openpyxl.load_workbook(excel_filename)
    for sheet in workbook:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{sheet.max_row}"
    workbook.save(excel_filename)
    print(f"{excel_filename} 已格式化")

# 主流程
def main():
    # 数据库连接信息
    host = '18.178.159.230'
    port = 3366
    user = 'bigdata'
    password = 'uvb5SOSmLH8sCoSU'
    database = 'finance_1000'
    table_name = 'finance_pay_records'
    use_fetch_member_ids = False

    start_time = datetime.datetime.now()
    print(f"开始: {start_time.strftime('%Y-%m-%d %H:%M')}")

    engine = get_db_engine(host, port, user, password, database)
    if not engine:
        return

    try:
        member_ids = fetch_member_ids(engine, database, table_name) if use_fetch_member_ids else ""
        chunks = fetch_main_data(engine, member_ids, database, table_name)
        if chunks is None:
            return

        excel_filename = generate_excel_filename(database, table_name)
        total_rows = 0
        tasks = []

        # 准备多进程任务
        for i, df_chunk in enumerate(tqdm(chunks, desc="准备数据")):
            sheet_name = f'Sheet{i + 1}'
            tasks.append((df_chunk, sheet_name, excel_filename))

        # 使用多进程写入 Excel
        with Pool(processes=cpu_count() - 1) as pool:
            total_rows = sum(pool.map(write_chunk_to_excel, tasks))

        # 获取列数
        sample_sql = f"SELECT * FROM {database}.{table_name} LIMIT 1"
        column_count = len(pd.read_sql(sample_sql, engine).columns)

        # 格式化 Excel
        format_excel(excel_filename, column_count)
        print(f"导出 {total_rows} 行到 {excel_filename}")

    finally:
        engine.dispose()  # 释放数据库连接池

    print(f"耗时: {(datetime.datetime.now() - start_time).total_seconds():.2f} 秒")

if __name__ == "__main__":
    main()
