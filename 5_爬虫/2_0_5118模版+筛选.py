import os
import zipfile
import io
import pandas as pd
import chardet
import requests
from bs4 import BeautifulSoup
import whois
import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from multiprocessing import Pool, cpu_count

# 工具函数
def detect_encoding(byte_data):
   result = chardet.detect(byte_data)
   return result['encoding'] if result['confidence'] > 0.5 else 'utf-8'

def extract_filename_info(file_name):
   match = re.search(r'5118-(.*?)(?=行业代表网站域名链接_)', file_name)
   return match.group(1) if match else None

def format_url(url):
   return f"http://{url}" if not url.startswith(("http://", "https://")) else url

def check_responsive(url):
   try:
       soup = BeautifulSoup(requests.get(format_url(url), timeout=5).text, 'html.parser')
       return "响应式网站" if soup.find('meta', attrs={'name': 'viewport'}) else "非响应式网站"
   except Exception as e:
       return f"无法访问: {e}"

def get_domain_last_updated(url):
   try:
       domain = url.split("//")[-1].split("/")[0]
       updated_date = whois.whois(domain).updated_date
       if isinstance(updated_date, list):
           updated_date = updated_date[0]
       return updated_date.replace(tzinfo=None) if updated_date and hasattr(updated_date, 'tzinfo') else updated_date
   except Exception:
       return None

def get_brand_terms(website_name):
   try:
       response = requests.get(format_url(website_name), timeout=5)
       response.encoding = response.apparent_encoding
       title = BeautifulSoup(response.text, 'html.parser').find('title')
       if title and title.text.strip():
           return ', '.join(filter(None, re.split(r'[\s\W_]+', title.text.strip())))
       domain = website_name.split("//")[-1].split("/")[0].replace('www.', '').split('.')[0]
       return domain or "未识别品牌词"
   except Exception:
       return ""

# 并行处理单行数据
def process_row(row_data):
   row = pd.Series(row_data)
   return pd.Series({
       **row,
       '响应式检测': check_responsive(row['网站名称']),
       '域名更新时间': get_domain_last_updated(row['网站名称']),
       '品牌词数据': get_brand_terms(row['网站名称'])
   })

def read_zip_csv_skip_first_row(folder_path):
   dfs = []
   for file_name in os.listdir(folder_path):
       if file_name.endswith('.zip'):
           file_info = extract_filename_info(file_name)
           with zipfile.ZipFile(os.path.join(folder_path, file_name), 'r') as zip_ref:
               for csv_file_name in zip_ref.namelist():
                   if csv_file_name.endswith(('txt', 'csv')):
                       with zip_ref.open(csv_file_name) as csv_file:
                           raw_data = csv_file.read()
                           encoding = detect_encoding(raw_data)
                           df = pd.read_csv(io.StringIO(raw_data.decode(encoding, errors='ignore')), skiprows=1)
                       if '网站名称' not in df.columns:
                           continue
                       if file_info:
                           df.insert(0, '来源文件', file_info)
                       dfs.append(df)
   return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

def save_to_excel(df, output_path, append=False):
   if append and os.path.exists(output_path):
       book = load_workbook(output_path)
       writer = pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
       df.to_excel(writer, sheet_name='Sheet1', startrow=book['Sheet1'].max_row, index=False, header=False)
   else:
       writer = pd.ExcelWriter(output_path, engine='openpyxl')
       df.to_excel(writer, sheet_name='Sheet1', index=False)
       ws = writer.sheets['Sheet1']
       ws.freeze_panes = 'A2'
       ws.auto_filter.ref = ws.dimensions
       col_idx = df.columns.get_loc('域名更新时间') + 1
       for row in range(2, ws.max_row + 1):
           ws[f'{get_column_letter(col_idx)}{row}'].number_format = 'YYYY-MM-DD'
   writer.close()

if __name__ == '__main__':
   folder_path = r"C:\Henvita\域名包"
   script_name = os.path.splitext(os.path.basename(__file__))[0]
   # 修复 f-string 语法问题
   current_time = datetime.datetime.now().strftime('%Y-%m-%d_%H.%M')
   output_path = f"{script_name}_{current_time}.xlsx"

   start_time = datetime.datetime.now()
   print(f"运行开始时间: {start_time.strftime('%Y-%m-%d %H:%M')}")

   # 读取数据
   data = read_zip_csv_skip_first_row(folder_path)
   if data.empty:
       print("无有效数据")
       exit()

   # 数据预处理
   data['网站名称'] = data['网站名称'].str.replace('www.', '', regex=False).str.lstrip('m.')
   data['辅助列'] = data['网站名称'].str.replace('.', '', regex=False)
   data['字符长度差'] = data['网站名称'].str.len() - data['辅助列'].str.len()
   data = pd.concat([
       data[data['字符长度差'] == 1],
       data[(data['字符长度差'] == 2) & (data['网站名称'].str.endswith('.com.cn') | data['网站名称'].str.startswith('m.'))]
   ], ignore_index=True)

   # 并行处理
   with Pool(processes=cpu_count()) as pool:
       processed_rows = list(tqdm(pool.imap(process_row, [row for _, row in data.iterrows()]), total=len(data), desc="处理进度"))
   processed_data = pd.concat([pd.DataFrame([row]) for row in processed_rows], ignore_index=True)

   # 一次性保存
   print(f"\n已处理 {len(processed_data)}/{len(processed_data)} 条数据:")
   print(processed_data.tail())
   save_to_excel(processed_data, output_path, append=False)

   end_time = datetime.datetime.now()
   print(f"运行结束时间: {end_time.strftime('%Y-%m-%d %H:%M')}")
   print("处理完成，数据已导出到:", output_path)
