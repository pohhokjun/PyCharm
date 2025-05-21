import pandas as pd
import datetime
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from multiprocessing import Pool
import logging
import pymysql
import multiprocessing
import uuid

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 数据库连接配置（支持环境变量，增强安全性）
CONFIG = {
    'host': os.getenv('DB_HOST', '18.178.159.230'),
    'port': int(os.getenv('DB_PORT', 3366)),
    'user': os.getenv('DB_USER', 'bigdata'),
    'password': os.getenv('DB_PASSWORD', 'uvb5SOSmLH8sCoSU')
}

# 报表配置
REPORTS = {
    '人数': {'db': 'bigdata', 'table': 'platform_daily_report', 'date_as_text': True},
    '金额': {'db': 'bigdata', 'table': 'platform_daily_report', 'date_as_text': True},
    '留存': {'db': 'bigdata', 'table': 'member_daily_statics', 'date_as_text': True},
    '存款': {'db': 'finance_1000', 'table': 'finance_pay_records', 'date_as_text': True},
    '取款': {'db': 'finance_1000', 'table': 'finance_withdraw', 'date_as_text': True},
    '红利': {'db': 'activity_1000', 'table': 'member_dividend', 'date_as_text': True}
}

# 站点ID到站点名称的映射
SITE_MAPPING = {
    1000: '好博体育',
    2000: '黄金体育',
    3000: '宾利体育',
    4000: 'HOME体育',
    5000: '亚洲之星',
    6000: '玖博体育',
    7000: '蓝火体育',
    8000: 'A7体育',
    9000: '皇马体育',
    9001: '摩根体育',
    9002: '友博体育'
}

# SQL 查询函数
def get_platform_report_sql(db, table):
    return f"""
    SELECT 
        site_id AS 站点, 
        DATE_FORMAT(statics_date, '%%Y-%%m-%%d') AS 日期,
        SUM(register_member_count) AS 注册人数,
        SUM(first_recharge_member_count) AS 首存人数,
        SUM(recharge_member_count) AS 充值人数,
        SUM(drawing_member_count) AS 取款人数,
        SUM(bet_member_count_settle) AS 投注人数
    FROM {db}.{table}
    WHERE statics_date >= DATE_SUB(CURDATE(), INTERVAL 31 DAY) 
      AND statics_date < CURDATE()
    GROUP BY site_id, statics_date
    ORDER BY site_id ASC, statics_date ASC
    """

def get_amount_report_sql(db, table):
    return f"""
    SELECT 
        site_id AS 站点, 
        DATE_FORMAT(statics_date, '%%Y-%%m-%%d') AS 日期,
        SUM(first_recharge_amount) AS 首充金额,
        SUM(recharge_amount) AS 存款金额,
        SUM(drawing_amount) AS 取款金额,
        SUM(valid_bet_amount_settle) AS 投注金额,
        (SUM(net_amount_settle) + SUM(early_settle_net_amount_settle)) AS 公司输赢,
        SUM(deposit_adjust_amount) AS 账户调整,
        -SUM(dividend_amount) AS 红利,
        -SUM(rebate_amount) AS 返水,
        -SUM(per_commission_amount) AS 代理佣金,
        (
        -(SUM(net_amount_settle) + SUM(early_settle_net_amount_settle) + SUM(deposit_adjust_amount)) * (MAX(group_amount_ratio) / 100)
        ) AS 集团分成,
        (
            SUM(net_amount_settle) + 
            SUM(early_settle_net_amount_settle) + 
            SUM(deposit_adjust_amount) - 
            SUM(dividend_amount) - 
            SUM(rebate_amount) - 
            SUM(per_commission_amount) - 
            ((SUM(net_amount_settle) + SUM(early_settle_net_amount_settle) + SUM(deposit_adjust_amount)) * (MAX(group_amount_ratio) / 100))
        ) AS 公司净收入
    FROM {db}.{table}
    WHERE statics_date >= DATE_SUB(CURDATE(), INTERVAL 31 DAY) 
      AND statics_date < CURDATE()
    GROUP BY site_id, statics_date
    ORDER BY site_id ASC, statics_date ASC
    """

def get_retention_report_sql(db, table):
    return f"""
    WITH date_range AS (
        SELECT DATE_SUB(DATE(CURDATE()), INTERVAL (n + 1) DAY) AS statics_date
        FROM (SELECT ROW_NUMBER() OVER () - 1 AS n 
              FROM information_schema.columns LIMIT 31) t
    ),
    base_data AS (
        SELECT site_id,
               DATE(first_deposit_time) AS cohort_date,
               member_id,
               bets,
               DATE(statics_date) AS activity_date
        FROM {db}.{table}
        WHERE first_deposit_time BETWEEN DATE_SUB(CURDATE(), INTERVAL 61 DAY) AND CURDATE()
          AND statics_date BETWEEN DATE_SUB(CURDATE(), INTERVAL 61 DAY) AND CURDATE()
    ),
    new_users AS (
        SELECT site_id, cohort_date, COUNT(DISTINCT member_id) AS 首存人数
        FROM base_data
        WHERE bets > 0
        GROUP BY site_id, cohort_date
    ),
    retention AS (
        SELECT site_id,
               activity_date AS statics_date,
               COUNT(DISTINCT CASE WHEN cohort_date = DATE_SUB(activity_date, INTERVAL 2 DAY)
                    AND bets > 0 THEN member_id END) AS `3日留存人数`,
               COUNT(DISTINCT CASE WHEN cohort_date = DATE_SUB(activity_date, INTERVAL 6 DAY)
                    AND bets > 0 THEN member_id END) AS `7日留存人数`,
               COUNT(DISTINCT CASE WHEN cohort_date = DATE_SUB(activity_date, INTERVAL 14 DAY)
                    AND bets > 0 THEN member_id END) AS `15日留存人数`,
               COUNT(DISTINCT CASE WHEN cohort_date = DATE_SUB(activity_date, INTERVAL 29 DAY)
                    AND bets > 0 THEN member_id END) AS `30日留存人数`
        FROM base_data
        GROUP BY site_id, activity_date
    )
    SELECT 
        COALESCE(n.site_id, r.site_id) AS 站点,
        DATE_FORMAT(d.statics_date, '%%Y-%%m-%%d') AS 日期,
        COALESCE(n.首存人数, 0) AS 首存人数,
        COALESCE(r.`3日留存人数`, 0) AS `3日留存人数`,
        COALESCE(r.`7日留存人数`, 0) AS `7日留存人数`,
        COALESCE(r.`15日留存人数`, 0) AS `15日留存人数`,
        COALESCE(r.`30日留存人数`, 0) AS `30日留存人数`
    FROM date_range d
    LEFT JOIN new_users n ON d.statics_date = n.cohort_date
    LEFT JOIN retention r ON d.statics_date = r.statics_date 
        AND (n.site_id = r.site_id OR (n.site_id IS NULL AND r.site_id IS NULL))
    ORDER BY 站点, d.statics_date ASC
    """

def get_payment_report_sql(db, table):
   return f"""
   WITH date_buckets AS (
       SELECT 
           site_id,
           pay_type,
           pay_status,
           order_amount,
           created_at,
           confirm_at,
           CASE 
               WHEN confirm_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY) THEN '近7日'
               WHEN confirm_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN '近30日'
           END AS 时间段
       FROM {db}.{table}
       WHERE category = 1
         AND pay_status IN (2, 4)
         AND confirm_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
         AND confirm_at <= CURDATE()
   )
   SELECT 
       f.site_id AS 站点,
       COALESCE(sv.dict_value, f.pay_type) AS 存款类型,
       SUM(CASE WHEN 时间段 = '近7日' THEN 1 ELSE 0 END) AS `7日订单数`,
       SUM(CASE WHEN 时间段 = '近7日' AND pay_status = 2 THEN 1 ELSE 0 END) AS `7日成功数量`,
       SUM(CASE WHEN 时间段 = '近7日' THEN order_amount ELSE 0 END) AS `7日订单金额`,
       SUM(CASE WHEN 时间段 = '近7日' AND pay_status = 2 THEN order_amount ELSE 0 END) AS `7日成功金额`,
       SUM(CASE WHEN 时间段 = '近30日' THEN 1 ELSE 0 END) AS `30日订单数`,
       SUM(CASE WHEN 时间段 = '近30日' AND pay_status = 2 THEN 1 ELSE 0 END) AS `30日成功数量`,
       SUM(CASE WHEN 时间段 = '近30日' THEN order_amount ELSE 0 END) AS `30日订单金额`,
       SUM(CASE WHEN 时间段 = '近30日' AND pay_status = 2 THEN order_amount ELSE 0 END) AS `30日成功金额`
   FROM date_buckets f
   LEFT JOIN (
       SELECT code, dict_value
       FROM control_1000.sys_dict_value
       WHERE initial_flag = 0
       AND dict_code = 'payment_method'
   ) sv ON f.pay_type = sv.code
   WHERE 时间段 IS NOT NULL
   GROUP BY f.site_id, f.pay_type, sv.dict_value
   ORDER BY f.site_id ASC
   """

def get_withdraw_report_sql(db, table):
   return f"""
   WITH date_buckets AS (
       SELECT 
           site_id,
           withdraw_type,
           draw_status,
           amount,
           created_at,
           confirm_at,
           CASE 
               WHEN confirm_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY) THEN '近7日'
               WHEN confirm_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY) THEN '近30日'
           END AS 时间段
       FROM {db}.{table}
       WHERE category = 1
         AND draw_status IN (402, 501)
         AND confirm_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
         AND confirm_at <= CURDATE()
   )
   SELECT 
       fw.site_id AS 站点,
       COALESCE(sv.dict_value, fw.withdraw_type) AS 取款类型,
       SUM(CASE WHEN 时间段 = '近7日' THEN 1 ELSE 0 END) AS `7日订单数`,
       SUM(CASE WHEN 时间段 = '近7日' AND draw_status = 402 THEN 1 ELSE 0 END) AS `7日成功数量`,
       SUM(CASE WHEN 时间段 = '近7日' THEN amount ELSE 0 END) AS `7日订单金额`,
       SUM(CASE WHEN 时间段 = '近7日' AND draw_status = 402 THEN amount ELSE 0 END) AS `7日成功金额`,
       SUM(CASE WHEN 时间段 = '近30日' THEN 1 ELSE 0 END) AS `30日订单数`,
       SUM(CASE WHEN 时间段 = '近30日' AND draw_status = 402 THEN 1 ELSE 0 END) AS `30日成功数量`,
       SUM(CASE WHEN 时间段 = '近30日' THEN amount ELSE 0 END) AS `30日订单金额`,
       SUM(CASE WHEN 时间段 = '近30日' AND draw_status = 402 THEN amount ELSE 0 END) AS `30日成功金额`
   FROM date_buckets fw
   LEFT JOIN (
       SELECT code, dict_value
       FROM control_1000.sys_dict_value
       WHERE initial_flag = 0
       AND dict_code IN ('payment_method', 'withdraw_type')
   ) sv ON fw.withdraw_type = sv.code
   WHERE 时间段 IS NOT NULL
   GROUP BY fw.site_id, fw.withdraw_type, sv.dict_value
   ORDER BY fw.site_id ASC
   """

def get_dividend_report_sql(db, table):
    return f"""
    SELECT 
        a1_md.site_id AS 站点,
        DATE_FORMAT(a1_md.updated_at, '%%Y-%%m-%%d') AS 日期,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '升级红利' THEN 1 ELSE 0 END) AS 升级红利_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '升级红利' THEN a1_md.money ELSE 0 END) AS 升级红利_金额,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '存款优惠' THEN 1 ELSE 0 END) AS 存款优惠_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '存款优惠' THEN a1_md.money ELSE 0 END) AS 存款优惠_金额,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '每周红利' THEN 1 ELSE 0 END) AS 每周红利_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '每周红利' THEN a1_md.money ELSE 0 END) AS 每周红利_金额,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '活动红利' THEN 1 ELSE 0 END) AS 活动红利_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '活动红利' THEN a1_md.money ELSE 0 END) AS 活动红利_金额,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '豪礼红利' THEN 1 ELSE 0 END) AS 豪礼红利_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) = '豪礼红利' THEN a1_md.money ELSE 0 END) AS 豪礼红利_金额,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) IN ('代理红利', '代理红利(代理后台-推广红利)') THEN 1 ELSE 0 END) AS 代理红利_人数,
        SUM(CASE WHEN COALESCE(sv_bonus.dict_value, a1_md.category) IN ('代理红利', '代理红利(代理后台-推广红利)') THEN a1_md.money ELSE 0 END) AS 代理红利_金额
    FROM activity_1000.member_dividend a1_md
    LEFT JOIN activity_1000.operation_activity_info a1_oci 
        ON a1_md.activity_id = a1_oci.id
    LEFT JOIN (
        SELECT code, dict_value
        FROM control_1000.sys_dict_value
        WHERE dict_code = 'bonus_type'
    ) sv_bonus ON a1_md.category = sv_bonus.code
    WHERE a1_md.status = 2
        AND a1_md.category NOT IN (999555)
        AND a1_md.created_at >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)
        AND a1_md.created_at <= CURDATE()
    GROUP BY a1_md.site_id, 
             DATE_FORMAT(a1_md.updated_at, '%%Y-%%m-%%d')
    """

# SQL 函数映射
SQL_FUNCTIONS = {
    '人数': get_platform_report_sql,
    '金额': get_amount_report_sql,
    '留存': get_retention_report_sql,
    '存款': get_payment_report_sql,
    '取款': get_withdraw_report_sql,
    '红利': get_dividend_report_sql
}


def process_report(args):
    """处理单个报表的查询和返回结果"""
    sheet_name, config, report_config = args
    try:
        engine = create_engine(
            f"mysql+pymysql://{config['user']}:{config['password']}@{config['host']}:{config['port']}/",
            pool_size=5,
            max_overflow=10,
            pool_timeout=30
        )
        try:
            sql = SQL_FUNCTIONS[sheet_name](report_config['db'], report_config['table'])
            df_iterator = pd.read_sql(sql, engine, chunksize=50000)
            data = []
            columns = None
            for chunk in df_iterator:
                if columns is None:
                    columns = chunk.columns.tolist()
                # 替换站点ID为站点名称
                if '站点' in chunk.columns:
                    chunk['站点'] = chunk['站点'].map(SITE_MAPPING).fillna(chunk['站点'])
                # 识别数值列，处理小数或整数
                for col in chunk.select_dtypes(include=['float64', 'int64']).columns:
                    if '成功率' in col:  # 保留包含“成功率”的字段为两位小数
                        chunk[col] = chunk[col].round(2)
                    else:  # 其他数值列转为整数
                        chunk[col] = chunk[col].round(0).astype('Int64')
                data.extend(chunk.values.tolist())
            return sheet_name, columns, data, True
        finally:
            engine.dispose()
    except Exception as e:
        logging.error(f"处理 {sheet_name} 出错: {e}")
        return sheet_name, None, None, False

def write_to_excel(results, filename):
    """将结果写入 Excel 文件"""
    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    for sheet_name, columns, data, success in results:
        if not success or not columns:
            logging.warning(f"{sheet_name} 无数据或查询失败，跳过写入")
            continue
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(columns)
        for row in data:
            worksheet.append(row)
        worksheet.freeze_panes = 'A2'
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{worksheet.max_row}"

    workbook.save(filename)
    logging.info(f"数据已导出到 {filename}")

def main():
    start_time = datetime.datetime.now()
    logging.info(f"开始: {start_time.strftime('%Y-%m-%d %H:%M')}")

    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_filename = "0_Ultimate_Analysis.xlsx"

    pool_size = min(len(REPORTS), multiprocessing.cpu_count())
    with Pool(processes=pool_size) as pool:
        tasks = [(name, CONFIG, REPORTS[name]) for name in REPORTS]
        results = pool.map(process_report, tasks)

    write_to_excel(results, excel_filename)

    end_time = datetime.datetime.now()
    logging.info(f"结束: {end_time.strftime('%Y-%m-%d %H:%M')}, 耗时: {(end_time - start_time).total_seconds():.2f} 秒")

if __name__ == "__main__":
    main()

