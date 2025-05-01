
import os
import asyncio
import datetime
import pymysql
import pandas as pd
from telegram import Bot
from telegram.error import TelegramError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

# 配置常量
FOLDER_PATH = 'C:/Henvita/1_定时注单导出/收费站'
TELEGRAM_BOT_TOKEN = '7750313084:AAGci5ANeeyEacKJUESQuDHYyy8tLdl9m7Q'
CHAT_ID = '-1002415614868'
DB_CONFIG = {
   'host': '18.178.159.230',
   'port': 3366,
   'user': 'bigdata',
   'password': 'uvb5SOSmLH8sCoSU',
   'database': 'finance_1000'
}
SITE_MAPPING = {
   1000: "好博体育", 2000: "黄金体育", 3000: "宾利体育", 4000: "HOME体育",
   5000: "亚洲之星", 6000: "玖博体育", 7000: "蓝火体育", 8000: "A7体育",
   9000: "K9体育", 9001: "摩根体育", 9002: "幸运体育"
}

def to_thousands(number):
   return f"{number:,.0f}"

def delete_files(directory):
   for filename in os.listdir(directory):
       file_path = os.path.join(directory, filename)
       if os.path.isfile(file_path):
           os.remove(file_path)
           print(f"删除文件: {file_path}")

def process_dataframe(df, period='daily'):
   group_key = '日期' if period == 'daily' else '月份'
   columns = ['id', group_key, '站点ID', '渠道类型', '存款调整上分金额', '存款调整下分金额',
              '充值金额', '提款金额', '注册人数', '首存人数', '首存金额', '首次有效充值会员数',
              '首次有效充值金额', '存款人数', '提款人数', '投注人数', '有效投注', '投注额',
              '公司输赢含提前结算', '红利', '返水', '个人佣金金额', '团队佣金金额',
              '提前结算净额', '账号调整', '首次充值注册比例', '人均首次充值', '存提差',
              '提款充值比例', '净投注金额比例', '公司输赢', '集团分成', '团队金额比例',
              '场馆费', '投注人数(结算)', '有效投注(结算)', '投注额(结算)',
              '公司输赢含提前结算(结算)', '提前结算(结算)', '公司输赢(结算)',
              '盈余比例(结算)', '场馆费(结算)', '打赏金额', '集团分成(结算)',
              '存款手续费', '提款手续费', '分成比例']
   df.columns = columns

   # 仅保留需要的字段
   required_cols = ['站点ID', group_key, '注册人数', '首存人数', '首存金额', '存款人数',
                   '充值金额', '提款人数', '提款金额', '存提差', '账号调整', '投注人数',
                   '投注额', '有效投注', '公司输赢含提前结算', '红利', '返水',
                   '个人佣金金额', '团队佣金金额']
   df = df[required_cols]

   agg_cols = {
       '注册人数': 'sum', '首存人数': 'sum', '首存金额': 'sum', '存款人数': 'sum',
       '充值金额': 'sum', '提款人数': 'sum', '提款金额': 'sum', '存提差': 'sum',
       '账号调整': 'sum', '投注人数': 'sum', '投注额': 'sum', '有效投注': 'sum',
       '公司输赢含提前结算': 'sum', '红利': 'sum', '返水': 'sum',
       '个人佣金金额': 'sum', '团队佣金金额': 'sum'
   }
   grouped = df.groupby(['站点ID', group_key])[list(agg_cols.keys())].sum().reset_index()

   # 计算集团分成、代理佣金、公司净收入
   grouped['集团分成'] = (grouped['公司输赢含提前结算'] + grouped['账号调整']) * 0.12
   grouped['代理佣金'] = grouped['个人佣金金额'] + grouped['团队佣金金额']
   grouped['公司净收入'] = (grouped['公司输赢含提前结算'] + grouped['账号调整'] -
                            grouped['红利'] - grouped['返水'] - grouped['代理佣金'] -
                            grouped['集团分成'])

   # 删除不需要的中间字段
   grouped.drop(columns=['个人佣金金额', '团队佣金金额'], inplace=True)

   # 过滤当前月份数据
   current_month = datetime.datetime.now().strftime("%Y-%m")
   grouped = grouped[grouped[group_key].str.startswith(current_month)]

   # 格式化数字为千位分隔（除了站点ID和日期/月份列）
   numeric_cols = ['注册人数', '首存人数', '首存金额', '存款人数', '充值金额',
                   '提款人数', '提款金额', '存提差', '账号调整', '投注人数',
                   '投注额', '有效投注', '公司输赢含提前结算', '红利', '返水',
                   '集团分成', '代理佣金', '公司净收入']
   for col in numeric_cols:
       grouped[col] = grouped[col].apply(to_thousands)

   # 映射站点ID为站点名称
   grouped['站点ID'] = grouped['站点ID'].map(SITE_MAPPING)

   # 重命名列以匹配要求
   grouped.columns = ['站点ID', group_key, '注册人数', '首存人数', '首存金额', '存款人数',
                      '存款金额', '提款人数', '提款金额', '存提差', '账户调整', '投注人数',
                      '投注额', '有效投注', '公司输赢含提前结算', '红利', '返水',
                      '集团分成', '代理佣金', '公司净收入']

   return grouped

async def job(bot):
   print(f"当前时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

   all_data = {}

   with pymysql.connect(**DB_CONFIG) as conn:
       # 处理月报数据
       df_monthly = pd.read_sql_query(f"SELECT * FROM bigdata.platform_month_report;", conn)
       monthly_data = process_dataframe(df_monthly, period='monthly')
       all_data['月报'] = monthly_data

       # 处理日报数据
       df_daily = pd.read_sql_query(f"SELECT * FROM bigdata.platform_daily_report;", conn)
       daily_data = process_dataframe(df_daily, period='daily')
       daily_grouped = daily_data.groupby('站点ID')

       for site_id, site_df in daily_grouped:
           site_name = SITE_MAPPING.get(site_id, str(site_id))
           all_data[site_name] = site_df

   file_name = f"{FOLDER_PATH}/平台报表-月日报表.xlsx"
   wb = Workbook()

   # 定义格式（对应原 xlsxwriter 的格式）
   header_fill = PatternFill(start_color="76933C", end_color="76933C", fill_type="solid")
   header_font = Font(color="FFFFFF", bold=True, size=12)
   first_col_fill = PatternFill(start_color="9AD000", end_color="9AD000", fill_type="solid")
   first_col_font = Font(color="FFFFFF", bold=True, size=12)
   even_row_fill = PatternFill(start_color="D7F8C4", end_color="D7F8C4", fill_type="solid")
   odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

   # 按照 SITE_MAPPING 的顺序生成 sheet
   ordered_sheets = ['月报'] + [SITE_MAPPING[site_id] for site_id in sorted(SITE_MAPPING.keys()) if SITE_MAPPING[site_id] in all_data]

   for sheet_name in ordered_sheets:
       data = all_data.get(sheet_name)
       if data is not None and not data.empty:
           ws = wb.create_sheet(title=sheet_name) if sheet_name != '月报' else wb.active
           ws.title = sheet_name

           # 写入表头
           for col_idx, column in enumerate(data.columns, 1):
               cell = ws.cell(row=1, column=col_idx)
               cell.value = column
               cell.fill = header_fill
               cell.font = header_font
               cell.alignment = Alignment(horizontal="center")

           # 写入数据
           for row_idx, row in enumerate(data.values, 2):
               for col_idx, value in enumerate(row, 1):
                   cell = ws.cell(row=row_idx, column=col_idx)
                   cell.value = value
                   # 首列格式
                   if col_idx == 1:
                       cell.fill = first_col_fill
                       cell.font = first_col_font
                   # 交替行颜色（从第2列开始）
                   else:
                       cell.alignment = Alignment(horizontal="right")
                       if row_idx % 2 == 0:
                           cell.fill = even_row_fill
                       else:
                           cell.fill = odd_row_fill

           # 模拟 Excel 双击自适应列宽
           for col_idx in range(1, len(data.columns) + 1):
               column_letter = get_column_letter(col_idx)
               max_length = 0
               for row in ws[column_letter]:
                   try:
                       cell_len = sum(1 if ord(c) < 128 else 2 for c in str(row.value)) if row.value else 0
                       max_length = max(max_length, cell_len)
                   except:
                       pass
               adjusted_width = max_length * 1.15
               ws.column_dimensions[column_letter].width = max(adjusted_width, 1)

           # 冻结首行
           ws.freeze_panes = 'A2'

           # 设置月报 sheet 标签颜色为黑色
           if sheet_name == '月报':
               ws.sheet_properties.tabColor = "000000"

   # 保存文件
   wb.save(file_name)
   print(f"已生成合并报表: {file_name}")

   try:
       with open(file_name, 'rb') as file:
           await bot.send_document(chat_id=CHAT_ID, document=file)
       print(f"文件已发送: {file_name}")
   except TelegramError as e:
       print(f"发送文件 {file_name} 时出错: {e}")
   finally:
       if os.path.exists(file_name):
           os.remove(file_name)
           print(f"删除文件: {file_name}")

async def main():
   bot = Bot(token=TELEGRAM_BOT_TOKEN)
   while True:
       await job(bot)
       await asyncio.sleep(3600)

if __name__ == "__main__":
   asyncio.run(main())

