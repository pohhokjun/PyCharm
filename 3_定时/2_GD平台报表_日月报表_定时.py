import os
import asyncio
import datetime
import pymysql
import pandas as pd
from telegram import Bot
from telegram.error import TelegramError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, numbers

# 配置常量
FOLDER_PATH = 'C:/Henvita/1_定时注单导出/收费站'
TELEGRAM_BOT_TOKEN = '7750313084:AAGci5ANeeyEacKJUESQuDHYyy8tLdl9m7Q'
CHAT_ID = '7523061850'
DB_CONFIG = {
   'host': '18.178.159.230',
   'port': 3366,
   'user': 'bigdata',
   'password': 'uvb5SOSmLH8sCoSU',
   'database': 'finance_1000'
}
SITE_MAPPING = {
   1000: "好博体育", 2000: "黄金体育", 3000: "宾利体育", 4000: "HOME体育",
   5000: "亚洲之星", 6000: "玖博体育", 7000: "蓝火体育", 8000: "A7体育",
   9000: "K9体育", 9001: "摩根体育", 9002: "友博体育"
}

def to_thousands(number):
   return f"{number:,.0f}"

def delete_files(directory):
   for filename in os.listdir(directory):
       file_path = os.path.join(directory, filename)
       if os.path.isfile(file_path):
           os.remove(file_path)
           print(f"删除文件: {file_path}")

async def job(bot):
   print(f"当前时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

   all_data = {}

   with pymysql.connect(**DB_CONFIG) as conn:
       # 处理月报数据（保持不变）
       monthly_query = """
       SELECT
           site_id AS 站点,
           statics_month AS 月份,
           FORMAT(SUM(register_member_count), 0) AS 注册人数,
           FORMAT(SUM(first_recharge_member_count), 0) AS 首存人数,
           FORMAT(SUM(first_recharge_amount), 0) AS 首存金额,
           FORMAT(SUM(recharge_member_count), 0) AS 存款人数,
           FORMAT(SUM(recharge_amount), 0) AS 存款金额,
           FORMAT(SUM(drawing_member_count), 0) AS 提款人数,
           FORMAT(SUM(drawing_amount), 0) AS 提款金额,
           FORMAT(SUM(recharge_drawing_sub), 0) AS 存提差,
           FORMAT(SUM(deposit_adjust_amount), 0) AS 账户调整,
           FORMAT(SUM(bet_member_count), 0) AS 投注人数,
           FORMAT(SUM(bet_amount), 0) AS 投注额,
           FORMAT(SUM(valid_bet_amount), 0) AS 有效投注,
           FORMAT(SUM(net_amount_settle), 0) AS 公司输赢含提前结算,
           FORMAT(SUM(dividend_amount), 0) AS 红利,
           FORMAT(SUM(rebate_amount), 0) AS 返水,
           FORMAT(SUM(net_amount_settle + deposit_adjust_amount) * 0.12, 0) AS 集团分成,
           FORMAT(SUM(per_commission_amount + team_commission_amount), 0) AS 代理佣金,
           FORMAT(
               SUM(net_amount_settle + deposit_adjust_amount - dividend_amount - rebate_amount - (per_commission_amount + team_commission_amount) - (net_amount_settle + deposit_adjust_amount) * 0.12),
               0
           ) AS 公司净收入
       FROM bigdata.platform_month_report
       WHERE statics_month LIKE CONCAT(DATE_FORMAT(CURDATE(), '%Y-%m'), '%')
       GROUP BY site_id, statics_month
       ORDER BY site_id;
       """
       df_monthly = pd.read_sql_query(monthly_query, conn)
       df_monthly['站点'] = df_monthly['站点'].map(SITE_MAPPING)
       all_data['月报'] = df_monthly

       # 处理日报数据
       daily_query = """
       SELECT
           site_id AS 站点,
           statics_date AS 日期,
           FORMAT(SUM(register_member_count), 0) AS 注册人数,
           FORMAT(SUM(first_recharge_member_count), 0) AS 首存人数,
           FORMAT(SUM(first_recharge_amount), 0) AS 首存金额,
           FORMAT(SUM(recharge_member_count), 0) AS 存款人数,
           FORMAT(SUM(recharge_amount), 0) AS 存款金额,
           FORMAT(SUM(drawing_member_count), 0) AS 提款人数,
           FORMAT(SUM(drawing_amount), 0) AS 提款金额,
           FORMAT(SUM(recharge_drawing_sub), 0) AS 存提差,
           FORMAT(SUM(deposit_adjust_amount), 0) AS 账户调整,
           FORMAT(SUM(bet_member_count), 0) AS 投注人数,
           FORMAT(SUM(bet_amount_settle), 0) AS 投注额,
           FORMAT(SUM(valid_bet_amount_settle), 0) AS 有效投注,
           FORMAT(SUM(net_amount_settle), 0) AS 公司输赢含提前结算,
           FORMAT(SUM(dividend_amount), 0) AS 红利,
           FORMAT(SUM(rebate_amount), 0) AS 返水,
           FORMAT(SUM(net_amount_settle + deposit_adjust_amount) * 0.12, 0) AS 集团分成,
           FORMAT(SUM(per_commission_amount + team_commission_amount), 0) AS 代理佣金,
           FORMAT(
               SUM(net_amount_settle + deposit_adjust_amount - dividend_amount - rebate_amount - (per_commission_amount + team_commission_amount) - (net_amount_settle + deposit_adjust_amount) * 0.12),
               0
           ) AS 公司净收入
       FROM bigdata.platform_daily_report
       WHERE statics_date LIKE CONCAT(DATE_FORMAT(CURDATE(), '%Y-%m'), '%')
       GROUP BY site_id, statics_date;
       """
       df_daily = pd.read_sql_query(daily_query, conn)
       df_daily['站点'] = df_daily['站点'].map(SITE_MAPPING)
       # 按日期升序排序
       df_daily = df_daily.sort_values(by='日期', ascending=True)
       daily_grouped = df_daily.groupby('站点')

       for site_id, site_df in daily_grouped:
           site_name = SITE_MAPPING.get(site_id, str(site_id))
           # 如果是“好博体育”，删除有效投注为0的行
           if site_name == "好博体育":
               site_df = site_df[site_df['有效投注'] != '0']
           all_data[site_name] = site_df

   file_name = f"{FOLDER_PATH}/平台报表-月日报表.xlsx"
   wb = Workbook()

   # 定义格式（对应原 xlsxwriter 的格式）
   header_fill = PatternFill(start_color="76933C", end_color="76933C", fill_type="solid")
   header_font = Font(color="FFFFFF", bold=True, size=12)
   first_col_fill = PatternFill(start_color="9AD000", end_color="9AD000", fill_type="solid")
   first_col_font = Font(color="FFFFFF", bold=True, size=12)
   even_row_fill = PatternFill(start_color="D7F8C4", end_color="D7F8C4", fill_type="solid")
   odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

   # 按照 SITE_MAPPING 的顺序生成 sheet
   ordered_sheets = ['月报'] + [SITE_MAPPING[site_id] for site_id in sorted(SITE_MAPPING.keys()) if SITE_MAPPING[site_id] in all_data]

   for sheet_name in ordered_sheets:
       data = all_data.get(sheet_name)
       if data is not None and not data.empty:
           ws = wb.create_sheet(title=sheet_name) if sheet_name != '月报' else wb.active
           ws.title = sheet_name

           # 写入表头
           for col_idx, column in enumerate(data.columns, 1):
               cell = ws.cell(row=1, column=col_idx)
               cell.value = column
               cell.fill = header_fill
               cell.font = header_font
               cell.alignment = Alignment(horizontal="center")

           # 写入数据
           for row_idx, row in enumerate(data.values, 2):
               for col_idx, value in enumerate(row, 1):
                   cell = ws.cell(row=row_idx, column=col_idx)
                   cell.value = value
                   # 首列格式
                   if col_idx == 1:
                       cell.fill = first_col_fill
                       cell.font = first_col_font
                   # 交替行颜色（从第2列开始）
                   else:
                       cell.alignment = Alignment(horizontal="right")
                       if row_idx % 2 == 0:
                           cell.fill = even_row_fill
                       else:
                           cell.fill = odd_row_fill

           # 设置日期列格式（第2列）
           for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
               for cell in row:
                   cell.number_format = 'yyyy-mm-dd'

           # 模拟 Excel 双击自适应列宽
           for col_idx in range(1, len(data.columns) + 1):
               column_letter = get_column_letter(col_idx)
               max_length = 0
               for row in ws[column_letter]:
                   try:
                       cell_len = sum(1 if ord(c) < 128 else 2 for c in str(row.value)) if row.value else 0
                       max_length = max(max_length, cell_len)
                   except:
                       pass
               adjusted_width = max_length * 1.15
               ws.column_dimensions[column_letter].width = max(adjusted_width, 1)

           # 冻结首行
           ws.freeze_panes = 'A2'

           # 设置月报 sheet 标签颜色为黑色
           if sheet_name == '月报':
               ws.sheet_properties.tabColor = "000000"

   # 保存文件
   wb.save(file_name)
   print(f"已生成合并报表: {file_name}")

   try:
       with open(file_name, 'rb') as file:
           await bot.send_document(chat_id=CHAT_ID, document=file)
       print(f"文件已发送: {file_name}")
   except TelegramError as e:
       print(f"发送文件 {file_name} 时出错: {e}")
   finally:
       if os.path.exists(file_name):
           os.remove(file_name)
           print(f"删除文件: {file_name}")

async def main():
   bot = Bot(token=TELEGRAM_BOT_TOKEN)
   while True:
       await job(bot)
       await asyncio.sleep(7200)

if __name__ == "__main__":
   asyncio.run(main())
