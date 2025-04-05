import os
import schedule
import time
from datetime import datetime, timedelta
import pymysql
import pandas as pd
import numpy as np
import shutil
from openpyxl import load_workbook
import xlwings as xw
import asyncio
from telegram import Bot
from telegram.error import TelegramError

# 你的模板文件路径
template_path = r"C:/Henvita/模版.xlsx"
folder_path = 'C:/Henvita/1_定时注单导出/收费站'
file_date = f'({datetime.now().strftime("%Y.%m.%d")})'

def to_thousands_separator(number):
    """
    将数字格式化为千位分隔符格式，保留0位小数。

    :param number: 要格式化的数字（整数或浮点数）
    :return: 千位分隔符格式的字符串
    """
    return f"{number:,.0f}"

def delete_all_files_in_directory(directory):
    if os.path.exists(directory) and os.path.isdir(directory):
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"删除文件: {file_path}")
    else:
        print("指定的目录不存在或不是一个目录。")

def calculate_totals(group_data):
    group_data = group_data.drop(columns=['站点ID'])
    sum_columns = ['注册数', '首存人数', '首存额', '存款人数', '取款人数', '存款额', '取款额', '存提差',
                   '投注人数', '投注额', '有效投注额', '公司输赢', '提前结算', '账户调整', '红利', '返水',
                   '代理佣金', '打赏收入', '集团分成', '公司净收入']
    sum_values = group_data[sum_columns].sum()
    total_row = pd.DataFrame(columns=group_data.columns)
    for col in sum_columns:
        total_row[col] = [sum_values[col]]

    total_row['转化率'] = np.where(total_row['注册数'] != 0,
                                   (total_row['首存人数'] / total_row['注册数']).apply(lambda x: '{:.2%}'.format(x)),
                                   '0.00%')
    total_row['人均首存'] = np.where(total_row['首存人数'] != 0,
                                     (total_row['首存额'] / total_row['首存人数']).round(2),
                                     0)
    total_row['提存率'] = np.where(total_row['存款额'] != 0,
                                   (total_row['取款额'] / total_row['存款额']).apply(lambda x: '{:.2%}'.format(x)),
                                   '0.00%')
    total_row['盈余比例'] = np.where(total_row['投注额'] != 0,
                                     ((total_row['公司输赢'] + total_row['提前结算']) / total_row['投注额']).apply(
                                         lambda x: '{:.2%}'.format(x)),
                                     '0.00%')
    total_row['集团分成比例'] = '12%'
    total_row['日期'] = '总计'
    result = pd.concat([group_data, total_row], ignore_index=True)
    return result

def process_dataframe(df, output_filename):
    output_path = os.path.join(folder_path, output_filename)
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["SheetJS"]
    start_row = 2
    start_col = 1
    for i, row_data in enumerate(df.values):
        for j, val in enumerate(row_data):
            ws.cell(row=start_row + i, column=start_col + j, value=val)

    app = xw.App(visible=False)
    try:
        temp_output_path = os.path.join(folder_path, "temp_" + output_filename)
        wb.save(temp_output_path)
        wb_xlwings = app.books.open(temp_output_path)
        ws_xlwings = wb_xlwings.sheets['SheetJS']

        total_row = None
        for row_num in range(1, ws_xlwings.api.UsedRange.Rows.Count + 1):
            if ws_xlwings.cells(row_num, 1).value == '总计':
                total_row = row_num
                break

        if total_row and total_row + 1 < 33:
            rows_to_delete = 33 - total_row
            ws_xlwings.api.Rows(f"{total_row + 1}:{33}").Delete()

        wb_xlwings.save(output_path)
        wb_xlwings.close()
        os.remove(temp_output_path)
    except Exception as e:
        print(f"处理 {output_filename} 时发生错误: {e}")
    finally:
        app.quit()

def job():
    print("当前时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    yesterday = datetime.now() - timedelta(days=1)
    current_month = yesterday.strftime('%Y-%m')
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    print(f'昨天日期:', yesterday_str)

    try:
        connection = pymysql.connect(
            host='18.178.159.230',
            port=3366,
            user='bigdata',
            password='uvb5SOSmLH8sCoSU',
            database='finance_1000'
        )
        query = "SELECT * FROM bigdata.platform_daily_report;"
        df = pd.read_sql_query(query, connection)
        connection.close()
    except pymysql.Error as e:
        print(f"数据库连接或查询错误: {e}")
        return

    df.columns = ['id', '日期', '站点ID', '渠道类型',
                  '存款调整上分金额', '存款调整下分金额',
                  '存款额', '取款额', '注册数',
                  '首存人数', '首存额',
                  '首次有效充值会员数', '首次有效充值金额',
                  '存款人数', '取款人数', '投注人数',
                  '有效投注', '投注额', '公司输赢含提前结算', '红利',
                  '返水', '个人佣金金额', '团队佣金金额',
                  '提前结算净额', '账户调整',
                  '首次充值注册比例', '人均首次充值',
                  '存提差', '提款充值比例',
                  '净投注金额比例', '公司输赢', '集团分成',
                  '团队金额比例', '场馆费', '投注人数(结算)',
                  '有效投注(结算)', '投注额(结算)', '公司输赢含提前结算(结算)',
                  '提前结算(结算)', '公司输赢(结算)',
                  '盈余比例(结算)', '场馆费(结算)', '打赏收入',
                  '集团分成(结算)', '存款手续费', '提款手续费', '分成比例']

    group_data = df.groupby(['站点ID', '日期']).agg({
        '注册数': 'sum',
        '首存人数': 'sum',
        '首存额': 'sum',
        '存款人数': 'sum',
        '存款额': 'sum',
        '取款人数': 'sum',
        '取款额': 'sum',
        '存提差': 'sum',
        '账户调整': 'sum',
        '投注人数(结算)': 'sum',
        '投注额(结算)': 'sum',
        '有效投注(结算)': 'sum',
        '公司输赢含提前结算(结算)': 'sum',
        '提前结算(结算)': 'sum',
        '红利': 'sum',
        '返水': 'sum',
        '个人佣金金额': 'sum',
        '团队佣金金额': 'sum',
        '打赏收入': 'sum'
    }).reset_index()

    group_data['转化率'] = np.where(group_data['注册数'] != 0,
                                    (group_data['首存人数'] / group_data['注册数']).apply(lambda x: '{:.2%}'.format(x)),
                                    '0.00%')
    group_data['人均首存'] = np.where(group_data['首存人数'] != 0,
                                      (group_data['首存额'] / group_data['首存人数']).round(2),
                                      0)
    group_data['提存率'] = np.where(group_data['存款额'] != 0,
                                    (group_data['取款额'] / group_data['存款额']).apply(lambda x: '{:.2%}'.format(x)),
                                    '0.00%')
    group_data['盈余比例'] = np.where(group_data['投注额(结算)'] != 0,
                                      (group_data['公司输赢含提前结算(结算)'] / group_data['投注额(结算)']).apply(
                                          lambda x: '{:.2%}'.format(x)),
                                      '0.00%')
    group_data['公司输赢'] = group_data['公司输赢含提前结算(结算)'] - group_data['提前结算(结算)']
    group_data['集团分成比例'] = '12%'
    group_data['集团分成(结算)'] = (group_data['公司输赢含提前结算(结算)'] + group_data['账户调整']) * 0.12
    group_data['代理佣金'] = group_data['个人佣金金额'] + group_data['团队佣金金额']
    group_data['公司净收入'] = group_data['公司输赢含提前结算(结算)'] + group_data['账户调整'] - group_data['红利'] - group_data['返水'] - group_data['代理佣金'] - group_data['集团分成(结算)']
    group_data = group_data.drop(columns=['个人佣金金额', '团队佣金金额'])
    group_data.columns = ['站点ID', '日期', '注册数', '首存人数', '首存额', '存款人数', '存款额', '取款人数', '取款额', '存提差',
                           '账户调整', '投注人数', '投注额', '有效投注额', '公司输赢含提前结算', '提前结算',
                           '红利', '返水', '打赏收入', '转化率', '人均首存', '提存率', '盈余比例', '公司输赢', '集团分成比例', '集团分成',
                           '代理佣金', '公司净收入']
    group_data = group_data[['站点ID', '日期', '注册数', '首存人数', '转化率', '首存额', '人均首存', '存款人数', '取款人数',
                             '存款额', '取款额', '存提差', '提存率', '投注人数', '投注额', '有效投注额', '公司输赢', '提前结算',
                             '盈余比例', '账户调整', '红利', '返水', '代理佣金', '打赏收入', '集团分成比例', '集团分成', '公司净收入']]

    group_data['红利'] = -group_data['红利']
    group_data['返水'] = -group_data['返水']
    group_data['集团分成'] = -group_data['集团分成']

    group_data = group_data[group_data['日期'].str.startswith(current_month)]
    group_data['日期'] = pd.to_datetime(group_data['日期'])
    group_data = group_data[group_data['日期'] <= pd.Timestamp(yesterday)]
    group_data['日期'] = group_data['日期'].dt.strftime('%Y-%m-%d')

    dict_data = {-1: '直客', 0: '普通代理', 1: '官方代理'}
    df['渠道类型'] = df['渠道类型'].replace(dict_data)

    site_ids = [1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000]
    site_names = {
        1000: "好博体育",
        2000: "黄金体育",
        3000: "宾利体育",
        4000: "HOME体育",
        5000: "亚洲之星",
        6000: "玖博体育",
        7000: "蓝火体育",
        8000: "A7体育"
    }
    dataframes = {}
    for site_id in site_ids:
        site_data = group_data[group_data['站点ID'] == site_id].copy()
        if not site_data.empty and site_id in site_names:
            dataframes[f"{site_names[site_id]}平台日报表{file_date}.xlsx"] = site_data

    for filename, df_to_process in dataframes.items():
        processed_df = calculate_totals(df_to_process.copy())
        process_dataframe(processed_df, filename)

if __name__ == '__main__':
    job()