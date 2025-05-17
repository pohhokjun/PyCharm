import asyncio
import logging
import re
import time
import pandas as pd
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import os
import tldextract
import json
import requests
import whois

# 关键词和域名列表
KEYWORDS = '''
959娱乐3.0最新版本更新内容
'''
keywords = [keyword.strip() for keyword in KEYWORDS.split('\n') if keyword.strip()]
domains = """
http://www.hmc.edu.cn/
"""
domain_list = [domain.strip() for domain in domains.split('\n') if domain.strip()]

# 配置日志
logging.basicConfig(filename='scraper.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

# 提取元数据
def extract_metadata(soup):
    title = soup.find('title').text.strip() if soup.find('title') else ''
    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
    keywords_content = keywords_meta['content'].strip() if keywords_meta else ''
    description_meta = soup.find('meta', attrs={'name': 'description'})
    description_content = description_meta['content'].strip() if description_meta else ''
    return {'title': title, 'keywords': keywords_content, 'description': description_content}

# 提取“大家还在搜”数据
def extract_related_searches(soup):
    related_div = soup.find('div', class_='c-title', string='大家还在搜')
    if related_div and related_div.next_sibling:
        data_tool = related_div.next_sibling.find('div', class_='sc-feedback')
        if data_tool and 'data-tool' in data_tool.attrs:
            try:
                tool_data = json.loads(data_tool['data-tool'].replace("'", '"'))
                related_words = tool_data.get('feedback', {}).get('suggest', {}).get('ext', {}).get('relation_words', '')
                return related_words
            except json.JSONDecodeError:
                return ''
    return ''

# 检测内容类型
def detect_content_type(url, soup):
    if "article" in url.lower() or soup.find('article'): return "文章"
    elif "product" in url.lower() or soup.find('div', class_=re.compile('product')): return "产品页面"
    elif "forum" in url.lower() or soup.find('div', class_=re.compile('post|thread')): return "论坛"
    elif "video" in url.lower() or soup.find('video'): return "视频"
    else: return "未知"

# 检测网站类型（新功能：区分单站、泛目录、泛站）
def detect_site_type(url, soup):
    try:
        parsed_url = tldextract.extract(url)
        path = url.split(parsed_url.registered_domain)[-1]
        # 泛目录：URL以 .html 结尾
        if path.lower().endswith('.html'):
            return "泛目录"
        # 泛站：子域名不是 www, m, 或空
        if parsed_url.subdomain and parsed_url.subdomain not in ['www', 'm']:
            return "泛站"
        # 单站：其他情况（包括 www, m, 或无子域名）
        return "单站"
    except Exception:
        return "未知"

# 检查是否为响应式网站
def check_responsive(url):
    try:
        soup = BeautifulSoup(requests.get(url, timeout=3).text, 'html.parser')
        return "响应式网站" if soup.find('meta', attrs={'name': 'viewport'}) else "非响应式网站"
    except Exception as e:
        return f"无法访问: {e}"

# 获取域名最后更新时间
def get_domain_last_updated(url):
    try:
        domain = url.split("//")[-1].split("/")[0]
        updated_date = whois.whois(domain).updated_date
        if isinstance(updated_date, list):
            updated_date = updated_date[0]
        return updated_date.replace(tzinfo=None) if updated_date and hasattr(updated_date, 'tzinfo') else updated_date
    except Exception:
        return None

# 获取品牌词
def get_brand_terms(website_name):
    try:
        response = requests.get(website_name, timeout=3)
        response.encoding = response.apparent_encoding
        title = BeautifulSoup(response.text, 'html.parser').find('title')
        if title and title.text.strip():
            return ', '.join(filter(None, re.split(r'[\s\W_]+', title.text.strip())))
        domain = website_name.split("//")[-1].split("/")[0].replace('www.', '').split('.')[0]
        return domain or "未识别品牌词"
    except Exception:
        return ""

# 提取总站
def extract_main_site(url):
    try:
        if url == '无链接':
            return ''
        # 提取“www.”后、首个“/”前的部分
        match = re.match(r'https?://(www\.)?([^/]+)', url)
        return match.group(2) if match else ''
    except Exception:
        return ''

# 搜索关键词
async def search_keyword(page, keyword, all_results):
    search_url = f"https://m.baidu.com/s?wd={keyword}"
    try:
        await page.goto(search_url, timeout=15000)
        await page.wait_for_selector("div.result", timeout=15000)
        html = await page.content()
        soup = BeautifulSoup(html, 'html.parser')
        related_searches = extract_related_searches(soup)
        results = soup.select('div.result')
        for result in results:
            data_log = result.get('data-log', '{}')
            link = eval(data_log.replace("'", '"')).get('mu', '无链接') if data_log else '无链接'
            title_element = result.select_one('h3') or result.select_one('a[class*="title"]')
            title_text = title_element.text.strip() if title_element else ''
            metadata, content_type, site_type = ({'title': '', 'keywords': '', 'description': ''}, '未知', '未知') if link == '无链接' else await fetch_page_data(page, link)
            uci = await calculate_page_uci(page, link)
            responsive_type = check_responsive(link) if link != '无链接' else "无链接"
            last_updated = get_domain_last_updated(link) if link != '无链接' else None
            brand_terms = get_brand_terms(link) if link != '无链接' else ""
            main_site = extract_main_site(link)  # 新增总站提取
            all_results.append({
                '百度引擎': '百度',
                '搜索关键词': keyword,
                '大家还在搜': related_searches,
                '内容类型': content_type,
                '网站类型': site_type,
                '响应类型': responsive_type,
                '最后更新时间': last_updated,
                '网址': link,
                '总站': main_site,  # 新增总站字段
                '品牌词': brand_terms,
                '页面标题': metadata['title'],
                '关键词': metadata['keywords'],
                '描述': metadata['description'],
                'UCI': uci
            })
    except Exception as e:
        logging.error(f"关键词 {keyword} 搜索失败：{e}")
    print(f"完成关键词：{keyword}")

# 获取页面数据（修改以返回网站类型）
async def fetch_page_data(page, link):
    try:
        await page.goto(link, timeout=5000)
        await page.wait_for_selector("body", timeout=5000)
        page_content = await page.content()
        page_soup = BeautifulSoup(page_content, 'html.parser')
        return extract_metadata(page_soup), detect_content_type(page.url, page_soup), detect_site_type(link, page_soup)
    except Exception as e:
        logging.error(f"链接 {link} 访问失败：{e}")
        return {'title': '', 'keywords': '', 'description': ''}, '未知', '未知'

# 获取域名数据
async def get_domain_data(page, domain):
    try:
        await page.goto(domain, timeout=5000)
        await page.wait_for_selector("body", timeout=5000)
        content = len(await page.content())
        links = len(await page.query_selector_all("a"))
        return content, links
    except Exception as e:
        logging.error(f"域名 {domain} 访问失败：{e}")
        return 0, 0

# 计算 UCI
def calculate_uci(cf, lf, df, uf):
    return min(100, round(0.4 * cf + 0.3 * lf + 0.2 * df + 0.1 * uf, 2))

# 计算页面UCI
async def calculate_page_uci(page, link):
    try:
        content, links = await get_domain_data(page, link)
        cf = min(100, content / 10000 * 100)
        lf = min(100, links / 50 * 100)
        df = 100 if tldextract.extract(link).registered_domain in ["baidu.com", "qq.com", "163.com", "sohu.com", "sina.com"] else 0
        uf = 50
        uci = calculate_uci(cf, lf, df, uf)
        return uci
    except Exception as e:
        logging.error(f"计算页面 {link} 的UCI失败：{e}")
        return 0

# 处理域名
async def process_domain(page, domain):
    content, links = await get_domain_data(page, domain)
    cf = min(100, content / 10000 * 100)
    lf = min(100, links / 50 * 100)
    df = 100 if tldextract.extract(domain).registered_domain in ["baidu.com", "qq.com", "163.com", "sohu.com", "sina.com"] else 0
    uf = 50
    uci = calculate_uci(cf, lf, df, uf)
    return {"域名": domain, "UCI": uci}

# 主函数
async def main():
    start_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行开始时间 {start_time}")
    logging.info(f"运行开始时间 {start_time}")

    all_results = []
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_file = f"{script_name}_{time.strftime('%Y-%m-%d_%H.%M')}.xlsx"

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False)
        page = await browser.new_page()

        # 处理关键词
        for i, keyword in enumerate(keywords, 1):
            await search_keyword(page, keyword, all_results)
            df = pd.DataFrame(all_results, columns=[
                '百度引擎', '搜索关键词', '大家还在搜', '内容类型', '网站类型',
                '响应类型', '最后更新时间', '网址', '总站', '品牌词',  # 添加总站列
                '页面标题', '关键词', '描述', 'UCI'
            ])
            df = df.replace(ILLEGAL_CHARACTERS_RE, "", regex=True)
            print(f"\n当前数据结果（前5条）：\n{df.tail(5).to_string()}")
            print(f"完成进度：{i}/{len(keywords)} 个关键词")
            if i == 1:
                df.to_excel(excel_file, index=False, sheet_name='KEYWORDS', engine='openpyxl')
            else:
                with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, index=False, sheet_name='KEYWORDS')
            print(f"结果已保存至 {excel_file} 的 KEYWORDS")

            # 同步处理“大家还在搜”到“大家还在搜”工作表
            wb = load_workbook(excel_file)
            ws_keywords = wb['KEYWORDS']
            ws_related = wb.create_sheet('大家还在搜') if '大家还在搜' not in wb else wb['大家还在搜']
            all_values = []
            related_col_idx = 3  # “大家还在搜”列索引（1-based）
            for row in ws_keywords.iter_rows(min_row=2, max_row=ws_keywords.max_row, min_col=related_col_idx, max_col=related_col_idx):
                for cell in row:
                    if cell.value and '&' in str(cell.value):
                        split_values = str(cell.value).split('&')
                        all_values.extend(split_values)
                    elif cell.value:
                        all_values.append(str(cell.value))
            unique_values = list(dict.fromkeys(all_values))
            ws_related.delete_rows(1, ws_related.max_row)
            for j, value in enumerate(unique_values, 1):
                ws_related[f'A{j}'].value = value
            wb.save(excel_file)
            print(f"“大家还在搜”数据已同步更新至 {excel_file} 的“大家还在搜”工作表")

        # 处理域名
        domain_results = [await process_domain(page, domain) for domain in domain_list]
        df_domains = pd.DataFrame(domain_results)
        df_domains = df_domains.replace(ILLEGAL_CHARACTERS_RE, "", regex=True)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_domains.to_excel(writer, index=False, sheet_name='domain')
        print(f"域名数据已保存至 {excel_file} 的 domain")

        await browser.close()

    # 格式化 Excel
    wb = load_workbook(excel_file)
    ws_keywords = wb['KEYWORDS']
    ws_domain = wb['domain']
    ws_keywords.freeze_panes, ws_domain.freeze_panes = 'A2', 'A2'
    ws_keywords.auto_filter.ref, ws_domain.auto_filter.ref = ws_keywords.dimensions, ws_domain.dimensions
    desc_col_idx = 11  # “描述”列索引（1-based）
    for row in ws_keywords.iter_rows(min_row=2, max_row=ws_keywords.max_row, min_col=desc_col_idx, max_col=desc_col_idx):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    wb.save(excel_file)
    print(f"Excel 文件已格式化")

    end_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行结束时间 {end_time}")
    logging.info(f"运行结束时间 {end_time}")

if __name__ == "__main__":
    asyncio.run(main())
