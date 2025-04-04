import asyncio
import logging
import re
import time
import pandas as pd
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import load_workbook
import os
import tldextract

# 关键词和域名列表
KEYWORDS = '''
918博天堂btt
918博天堂官方app
'''
keywords = [keyword.strip() for keyword in KEYWORDS.split('\n') if keyword.strip()]

domains = """
https://m.micro-maker.com/
http://www.aierxingz.com/
https://www.bet365.com/
https://1xbetind.in/
https://m.jslpsp.com/
https://www.ruimeidachuju.com/
"""
domain_list = [domain.strip() for domain in domains.split('\n') if domain.strip()]

# 配置日志
logging.basicConfig(filename='scraper.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

# 提取元数据
def extract_metadata(soup):
    title = soup.find('title').text.strip() if soup.find('title') else ''
    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
    keywords_content = keywords_meta['content'].strip() if keywords_meta else ''
    description_meta = soup.find('meta', attrs={'name': 'description'})
    description_content = description_meta['content'].strip() if description_meta else ''
    return {'title': title, 'keywords': keywords_content, 'description': description_content}

# 检测内容类型
def detect_content_type(url, soup):
    if "article" in url.lower() or soup.find('article'): return "文章"
    elif "product" in url.lower() or soup.find('div', class_=re.compile('product')): return "产品页面"
    elif "forum" in url.lower() or soup.find('div', class_=re.compile('post|thread')): return "论坛"
    elif "video" in url.lower() or soup.find('video'): return "视频"
    else: return "未知"

# 搜索关键词
async def search_keyword(page, keyword, all_results):
    search_url = f"https://m.baidu.com/s?wd={keyword}"
    try:
        await page.goto(search_url, timeout=30000)
        await page.wait_for_selector("div.result", timeout=30000)
        html = await page.content()
        soup = BeautifulSoup(html, 'html.parser')
        results = soup.select('div.result')[:5]
        for result in results:
            data_log = result.get('data-log', '{}')
            link = eval(data_log.replace("'", '"')).get('mu', '无链接') if data_log else '无链接'
            title_element = result.select_one('h3') or result.select_one('a[class*="title"]')
            title_text = title_element.text.strip() if title_element else ''
            metadata, content_type = ({'title': '', 'keywords': '', 'description': ''}, '未知') if link == '无链接' else await fetch_page_data(page, link)
            all_results.append({'搜索关键词': keyword, '内容类型': content_type, '网址': link, '页面标题': metadata['title'], '关键词': metadata['keywords'], '描述': metadata['description']})
    except Exception as e:
        logging.error(f"关键词 {keyword} 搜索失败：{e}")
    print(f"完成关键词：{keyword}")

# 获取页面数据
async def fetch_page_data(page, link):
    try:
        await page.goto(link, timeout=10000)
        await page.wait_for_selector("body", timeout=10000)
        page_content = await page.content()
        page_soup = BeautifulSoup(page_content, 'html.parser')
        return extract_metadata(page_soup), detect_content_type(page.url, page_soup)
    except Exception as e:
        logging.error(f"链接 {link} 访问失败：{e}")
        return {'title': '', 'keywords': '', 'description': ''}, '未知'

# 获取域名数据
async def get_domain_data(page, domain):
    try:
        await page.goto(domain, timeout=10000)
        await page.wait_for_selector("body", timeout=10000)
        content = len(await page.content())
        links = len(await page.query_selector_all("a"))
        return content, links
    except Exception as e:
        logging.error(f"域名 {domain} 访问失败：{e}")
        return 0, 0

# 计算 UCI
def calculate_uci(cf, lf, df, uf):
    return min(100, round(0.4 * cf + 0.3 * lf + 0.2 * df + 0.1 * uf, 2))

# 处理域名
async def process_domain(page, domain):
    content, links = await get_domain_data(page, domain)
    cf = min(100, content / 10000 * 100)
    lf = min(100, links / 50 * 100)
    df = 100 if tldextract.extract(domain).registered_domain in ["baidu.com", "qq.com", "163.com", "sohu.com", "sina.com"] else 0
    uf = 50
    uci = calculate_uci(cf, lf, df, uf)
    return {"域名": domain, "UCI": uci}

# 主函数
async def main():
    start_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行开始时间 {start_time}")
    logging.info(f"运行开始时间 {start_time}")

    all_results = []
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_file = f"{script_name}_{time.strftime('%Y-%m-%d_%H.%M')}.xlsx"

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False)
        page = await browser.new_page()

        # 处理关键词
        for i, keyword in enumerate(keywords):
            await search_keyword(page, keyword, all_results)
        df = pd.DataFrame(all_results)
        df.to_excel(excel_file, index=False, sheet_name='Sheet1', engine='openpyxl')
        print(f"已处理 {len(keywords)}/{len(keywords)} 个关键词，结果已保存至 {excel_file} 的 Sheet1")

        # 处理域名（在关键词任务完成后执行）
        domain_results = []
        for domain in domain_list:
            result = await process_domain(page, domain)
            domain_results.append(result)
        df_domains = pd.DataFrame(domain_results)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_domains.to_excel(writer, index=False, sheet_name='Sheet2')
        print(f"域名数据已保存至 {excel_file} 的 Sheet2")

        await browser.close()

    # 格式化 Excel
    wb = load_workbook(excel_file)
    ws1, ws2 = wb['Sheet1'], wb['Sheet2']
    ws1.freeze_panes, ws2.freeze_panes = 'A2', 'A2'
    ws1.auto_filter.ref, ws2.auto_filter.ref = ws1.dimensions, ws2.dimensions
    desc_col_idx = list(df.columns).index('描述') + 1
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=desc_col_idx, max_col=desc_col_idx):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')
    wb.save(excel_file)
    print(f"Excel 文件已格式化")

    end_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行结束时间 {end_time}")
    logging.info(f"运行结束时间 {end_time}")

if __name__ == "__main__":
    asyncio.run(main())