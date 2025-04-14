import socket
import subprocess
import logging
import time
import pandas as pd
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
import asyncio
from tqdm.asyncio import tqdm
import re
from openpyxl import load_workbook

logging.basicConfig(filename='ip_domain_query.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8', filemode='w')

DOMAIN_INPUT = """
wmihealth.com
17cn.net
"""
domains = [d.strip() for d in DOMAIN_INPUT.split('\n') if d.strip()]
if not domains:
    logging.error("域名列表为空，请检查输入")
    print("错误：域名列表为空")
    exit(1)

excel_filename = f"{__file__.split('.')[0]}_{time.strftime('%Y-%m-%d_%H.%M')}.xlsx"

async def get_ip_by_ping(domain):
    try:
        proc = await asyncio.create_subprocess_exec("ping", domain, "-n", "1", stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE, encoding='utf-8', errors='ignore')
        stdout, _ = await proc.communicate()
        match = re.search(r'\[(\d+\.\d+\.\d+\.\d+)\]', stdout)
        return match.group(1) if match else None
    except Exception as e:
        logging.error(f"Ping 获取 IP 失败 {domain}: {e}")
        return None

async def resolve_domain_ip(domain):
    try:
        ip = (await asyncio.get_event_loop().getaddrinfo(domain, None))[0][4][0]
    except socket.gaierror:
        logging.warning(f"{domain} 无法解析，尝试 Ping")
        ip = await get_ip_by_ping(domain)
    if ip:
        parts = ip.split('.')
        if len(parts) == 4:
            prefix = ".".join(parts[:3])
            num = int(parts[3])
            return [f"{prefix}.{i}" for i in range(num - 50, num + 51) if 0 <= i <= 255]
        return [ip]
    return ["解析失败"]

async def query_ip_domains(ip, page):
    if ip == "解析失败":
        return ["无法查询: IP解析失败"]
    try:
        await page.goto(f"https://ipchaxun.com/{ip}", timeout=5000)
        await page.wait_for_selector("body", timeout=5000)
        await asyncio.sleep(1)
        html = await page.content()
        soup = BeautifulSoup(html, 'html.parser')
        domains = [entry.text.strip() for entry in soup.select("span.date + a")]
        return domains if domains else ["未找到绑定域名"]
    except Exception as e:
        logging.error(f"查询失败 {ip}: {e}")
        return ["查询失败"]

async def get_meta_info(domain, page):
    for protocol in ["https", "http"]:
        url = f"{protocol}://{domain}"
        try:
            await page.goto(url, timeout=5000)
            await page.wait_for_selector("body", timeout=5000)
            html = await page.content()
            soup = BeautifulSoup(html, 'html.parser')
            title = soup.title.string.strip() if soup.title else ""
            keywords = next((meta.get('content', '').strip() for meta in soup.find_all('meta') if meta.get('name', '').lower() == 'keywords'), "")
            description = next((meta.get('content', '').strip() for meta in soup.find_all('meta') if meta.get('name', '').lower() == 'description'), "")
            if not any([title, keywords, description]):
                logging.warning(f"{domain} 无元信息，HTML: {html[:500]}")
            else:
                logging.info(f"{domain} - title: {title}, keywords: {keywords}, description: {description}")
            return title, keywords, description
        except Exception as e:
            logging.error(f"访问 {url} 失败: {e}")
            continue
    return "", "", ""

def save_to_excel(data, filename, mode='a'):
    df = pd.DataFrame(data).explode('绑定域名').drop_duplicates(subset=['域名', 'IP', '绑定域名'])
    with pd.ExcelWriter(filename, engine="openpyxl", mode=mode, if_sheet_exists="overlay" if mode == 'a' else None) as writer:
        startrow = writer.sheets['Sheet1'].max_row if mode == 'a' and 'Sheet1' in writer.sheets else 0
        df.to_excel(writer, index=False, header=startrow == 0, startrow=startrow)
    wb = load_workbook(filename)
    ws = wb.active
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(filename)

async def process_domain(domain):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--disable-gpu", "--no-sandbox", "--disable-dev-shm-usage", "--ignore-certificate-errors", "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"])
        page = await browser.new_page()
        try:
            ips = await resolve_domain_ip(domain)
            results = []
            for ip in ips:
                bound_domains = await query_ip_domains(ip, page)
                for bound_domain in bound_domains:
                    title, keywords, description = await get_meta_info(bound_domain, page)
                    result = {"域名": domain, "IP": ip, "绑定域名": bound_domain, "Title": title, "Keywords": keywords, "Description": description}
                    results.append(result)
                    print(f"域名：{domain}  IP：{ip}  绑定域名：{bound_domain}")
            return results
        finally:
            await browser.close()

async def main():
    start_time = time.strftime('%Y-%m-%d %H:%M')
    print(f"运行开始时间: {start_time}")
    total = len(domains)
    completed = 0
    for domain in tqdm(domains, desc="处理进度"):
        results = await process_domain(domain)
        completed += 1
        print(f"完成进度: {completed}/{total}")
        save_to_excel(results, excel_filename, mode='w' if completed == 1 else 'a')
    print(f"已保存 Excel: {excel_filename}")
    end_time = time.strftime('%Y-%m-%d %H:%M')
    print(f"运行结束时间: {end_time}")

if __name__ == "__main__":
    asyncio.run(main())
