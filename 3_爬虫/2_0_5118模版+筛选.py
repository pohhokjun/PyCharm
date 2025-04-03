import os
import zipfile
import io
import pandas as pd
import chardet
import requests
from bs4 import BeautifulSoup
import whois
import datetime
from tqdm import tqdm
from openpyxl import load_workbook
from multiprocessing import Pool
from pytrends.request import TrendReq
from openpyxl.utils import get_column_letter


def detect_encoding(byte_data):
    """检测文件编码"""
    result = chardet.detect(byte_data)
    return result['encoding'] if result['confidence'] > 0.5 else 'utf-8'


def extract_filename_info(file_name):
    """从文件名中提取 5118- 之后 和 行业代表网站域名链接_ 之前的内容"""
    start_idx = file_name.find("5118-") + len("5118-") if "5118-" in file_name else -1
    end_idx = file_name.find("行业代表网站域名链接_") if "行业代表网站域名链接_" in file_name else -1
    if start_idx != -1 and end_idx != -1 and start_idx < end_idx:
        return file_name[start_idx:end_idx]
    return None


def remove_str(string1, string2):
    return string1.replace(string2, '')


def left_trip(string, string2=None):
    return string.lstrip(string2)


def read_zip_csv_skip_first_row(folder_path):
    """读取 ZIP 文件中的 CSV，跳过第一行"""
    dfs = []
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.zip'):
            file_info = extract_filename_info(file_name)
            with zipfile.ZipFile(os.path.join(folder_path, file_name), 'r') as zip_ref:
                for csv_file_name in zip_ref.namelist():
                    if csv_file_name.endswith('txt') or csv_file_name.endswith('csv'):
                        with zip_ref.open(csv_file_name) as csv_file:
                            raw_data = csv_file.read()
                            encoding = detect_encoding(raw_data)
                            df = pd.read_csv(io.StringIO(raw_data.decode(encoding, errors='ignore')),
                                             encoding=encoding, skiprows=1)
                        if '网站名称' not in df.columns:
                            continue
                        if file_info:
                            df.insert(0, '来源文件', file_info)
                        dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def format_url(url):
    """确保 URL 有 http:// 或 https:// 前缀"""
    if not url.startswith("http://") and not url.startswith("https://"):
        return "http://" + url
    return url


def check_responsive(url):
    """检查网站是否为响应式设计"""
    formatted_url = format_url(url)
    try:
        response = requests.get(formatted_url, timeout=5)
        soup = BeautifulSoup(response.text, 'html.parser')
        if soup.find('meta', attrs={'name': 'viewport'}):
            return "响应式网站"
        else:
            return "非响应式网站"
    except Exception as e:
        return f"无法访问: {e}"


def get_domain_last_updated(url):
    """获取域名最后更新时间"""
    try:
        domain = url.split("//")[-1].split("/")[0]
        domain_info = whois.whois(domain)
        if isinstance(domain_info.updated_date, list):
            return domain_info.updated_date[0] if domain_info.updated_date else None
        return domain_info.updated_date if domain_info.updated_date else None
    except Exception as e:
        return None


def get_brand_terms(website_name):
    """获取网站名称的品牌词数据，基于域名推断"""
    try:
        # 从网站名称中提取核心品牌词（假设是域名部分）
        domain = website_name.split("//")[-1].split("/")[0] if "//" in website_name else website_name
        brand = domain.split('.')[0].replace('www', '').strip()  # 移除 'www' 和多余字符
        return brand if brand else "未识别品牌词"
    except Exception as e:
        return f"品牌词提取失败: {e}"


def process_row(row):
    """处理单行数据"""
    row_data = row.to_frame().T
    row_data['响应式检测'] = check_responsive(row['网站名称'])
    row_data['域名更新时间'] = get_domain_last_updated(row['网站名称'])
    row_data['品牌词数据'] = get_brand_terms(row['网站名称'])
    return row_data


def save_to_excel(df, output_path, append=False):
    """保存或追加数据到 Excel，冻结首行并设置筛选"""
    if append and os.path.exists(output_path):
        book = load_workbook(output_path)
        writer = pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        df.to_excel(writer, sheet_name='Sheet1', startrow=book['Sheet1'].max_row, index=False, header=False)
        writer.close()
    else:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            ws = writer.sheets['Sheet1']
            ws.freeze_panes = 'A2'  # 冻结首行
            ws.auto_filter.ref = ws.dimensions  # 设置筛选
            # 设置“域名更新时间”列为日期格式
            col_idx = df.columns.get_loc('域名更新时间') + 1
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                ws[f'{col_letter}{row}'].number_format = 'YYYY-MM-DD'


if __name__ == '__main__':
    folder_path = r"C:\Henvita\域名包"  # 你的 ZIP 文件所在路径
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    output_path = f"{script_name}_{datetime.datetime.now().strftime('%Y-%m-%d_%H.%M')}.xlsx"

    start_time = datetime.datetime.now()
    print(f"运行开始时间: {start_time.strftime('%Y-%m-%d %H:%M')}")

    # 读取 ZIP 数据并跳过第一行
    data = read_zip_csv_skip_first_row(folder_path)

    # 处理数据
    if not data.empty:
        data['网站名称'] = data['网站名称'].apply(lambda x: remove_str(x, 'www.') if isinstance(x, str) else x)
        data['辅助列'] = data['网站名称'].apply(lambda x: remove_str(x, '.') if isinstance(x, str) else x)
        data['字符长度差'] = data['网站名称'].str.len() - data['辅助列'].str.len()

        # 筛选符合条件的数据
        data1 = data[data['字符长度差'] == 1]
        data2 = data[(data['字符长度差'] == 2) & (data['网站名称'].str.endswith('.com.cn'))]
        data3 = data[(data['字符长度差'] == 2) & (data['网站名称'].str.startswith('m.'))]
        data = pd.concat([data1, data2, data3], ignore_index=True)
        data['网站名称'] = data['网站名称'].apply(lambda x: left_trip(x, 'm.') if isinstance(x, str) else x)

        # 分批处理并保存
        processed_data = pd.DataFrame()
        total_rows = len(data)
        for i, row in enumerate(tqdm(data.iterrows(), total=total_rows, desc="处理进度")):
            result = process_row(row[1])
            processed_data = pd.concat([processed_data, result], ignore_index=True)

            # 每处理5个或最后一行时打印并保存
            if (i + 1) % 5 == 0 or i == total_rows - 1:
                print(f"\n已处理 {i + 1}/{total_rows} 条数据:")
                print(processed_data.tail(5 if i + 1 >= 5 else i + 1))
                save_to_excel(processed_data, output_path, append=(i + 1) > 5)
                processed_data = pd.DataFrame()  # 清空已保存的数据

    end_time = datetime.datetime.now()
    print(f"运行结束时间: {end_time.strftime('%Y-%m-%d %H:%M')}")
    print("处理完成，数据已导出到:", output_path)