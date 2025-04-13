import asyncio
import logging
import re
import time
import pandas as pd
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE  # 导入 ILLEGAL_CHARACTERS_RE
import os
import tldextract
import json
import requests
import whois  # 新增 whois 库用于获取域名更新时间

# 关键词和域名列表
KEYWORDS = '''
菠菜棋牌注册入口
菠菜娱乐棋牌官网
菠菜娱乐平台登录入口
菠菜棋牌官方正版入口地址是什么
菠菜棋牌娱乐最新版本更新内容分享时间
芒果棋牌
神殿娱乐棋牌ios版
芒果娱乐棋牌官网版
芒果棋牌正版入口
迷路棋牌
芒果棋牌2023官方最新版
东篱棋牌
众乐棋牌
全盛棋牌69cc2.0.apk
芒果棋牌娱乐MEG
菠菜棋牌最新版2024
菠菜棋牌2.0.0版最新版本更新内容
菠菜棋牌官方正版
菠菜棋牌2024官方版
菠菜棋牌平台
菠菜棋牌官网登录入口
菠菜棋牌正版2024
菠菜棋牌正版2023年最新版特色功能
菠菜棋牌手机版免费
菠菜棋牌官方正版入口
途游德州官网下载
天天德州2025官方版
途游德州
德州app全没了
德州扑克牌(免费版)
天天欢乐德州app下载
《天天欢乐德州》官网
免费德州app游戏推荐
不赌钱的德州app
哪个游戏app能玩德州
天天德州官方版下载
天天德州免费版
菠菜棋牌亚洲版2024新版本介绍
菠菜棋牌亚洲版2024年最新版特色服务介绍
亚洲版菠菜棋牌官方版特色功能介绍
新用户放水棋牌
空间棋牌v6.1.0
pg棋牌网
棋牌加盟十大品牌
棋牌室最牛十大排名
棋牌平台十大排名
大菠萝棋牌官网版高清
国际棋牌
来乐棋牌
开瑞棋牌官网2022正版
果果棋牌
大棋牌
588棋牌正版官网
波克棋牌官方版
波克棋牌新版本
大菠萝棋牌官方网站
多乐棋牌
菠菜棋牌大厅最新版本更新内容
菠菜体育app官方最新版本更新内容
菠菜体育app下载安装
足彩胜负平推荐英冠前瞻
英国幸运5分彩
法甲买球app
线上正规买球
足球跟单计划资金表
2025年4月8日足球竞彩比赛
足球比分结果500
如何购买足彩彩票
足球数据
beat365在线官网登录入口
英国正版365官网
Bet356体育在线登录入口
365best体育官网登录入口
www.bet8338.com
菠菜真人娱乐是正规平台吗安全吗
菠菜真人娱乐app的最新版本更新时间
菠菜真人电视剧免费播放
菠菜真人免费追剧
菠菜真人2024年开播时间表
菠菜真人电视剧最新更新内容
菠菜真人2024年几月几号上映
菠菜真人2024年几月开播
菠菜真人集团背景有多强
菠菜真人集团创始人背景
菠菜真人集团最厉害三个股东
菠菜真人创始人简介
菠菜真人老板个人资料
蔬菜集团董事长
菠菜真人老板是谁
菠菜真人老板的背景资料
菠菜真人主页的访问历史和历史记录
菠菜真人主页的登录方式和历史记录
菠菜真人主页的登录方式和历史背景介绍
菠菜真人网页版登录入口
菠菜真人最新主页
菠菜真人竞猜app
菠菜真人竞猜app下载安装最新版
菠菜竞猜app官方最新版本更新内容
亚洲第一直播网
菠菜网投
菠菜资讯导航大全
亚洲菠菜公司排行
菠菜真人无删减版电视剧免费观看
菠菜真人视频免费观看高清版
菠菜真人无删减版在线播放
菠菜真人网页版登录入口最新版本更新内容
菠菜真人免费观看全部
菠菜真人高清免费观看
菠菜影院手机在线
菠菜真人无删减版大尺度电影叫什么名字
菠菜真人无删减版在线播放免费下载
菠菜真人无删减版大尺度2023
菠菜软件下载
菠菜天下app下载
bat365在线官网平台
亚傅网登录入口
亚博vip888官网网页登录
博乐体育官方入口
博乐官方下载
博乐体育官方平台登录
外网买球app十大平台
国外的合法彩票网站
国外彩票平台有哪些
外网竞彩预测
国外搏彩平台
国外足球彩票平台有哪些
国外体彩网站
国外football预测网站
国外足彩推荐网站
网赌刷返水5次了警察会知道吗
网赌刷返水5万判了几年
网赌刷流水5万判刑几年
网赌刷返水真赚钱吗
网赌一般查几年流水
网赌刷流水5次会坐牢吗
bob游戏平台官网
抓鱼游戏
bob官方手机网站
bob娱乐官网入口
bob电竞官网
bob官方网页版
bob游戏app
bob综合改成啥名了
bob游戏官方登录
771771威尼斯.cmApp
澳客官方app下载
澳客手机版app下载
74779·ccm
best365官方网站登录入口
365电子游戏官网入口
beat365·体育官网
全家福捕鱼高爆版
变态捕鱼高爆版百亿炮倍
好彩官网app
好彩捕鱼安装
好彩捕鱼1兆
博鱼彩票app官网入口下载
乐鱼app体育全站下载
博鱼体育综合APP下载
博鱼体彩
中国体育彩票app
中国体育彩票app官方下载
体育竞彩推荐平台
体彩足球竞彩平台
1999.c彩票网澳门
澳门官网彩票平台
2025澳门开奖结果记录
澳门官方彩票网版本2.4.6
澳彩官方正版app
中国福彩app免费下载
最全彩票app下载
彩票安卓版app下载
下载彩票app大全免费
785cc彩票app免费下载
竞彩中国体育彩票官网
首页竞彩网
竞彩足球官方网首页
足彩官网首页
中国体彩竞彩网官网首页
澳客竞彩足球官网
足彩竞彩首页官网
彩票竞彩网
中国体育竞彩足球彩票官网
中国竞彩网足彩网官网首页
亚洲彩票
BB贝博体彩平台下载
BB彩票首页入口
BB彩票官方网站
BB平台下载入口
BB彩票网登录入口
BB平台App
BB彩票下载
bb彩票平台大全
BB官方APP
BB彩票app官方网站入口
彩票在线试玩2000
凤凰新人注册送38元
大小单双彩票app平台官方
彩票app最新版
下载彩虹多多彩票安装
彩票诈骗的四个特征
彩票导师带计划群聊
彩票内部人员揭露彩票骗局
最新骗局彩票进精准群
LOL投注官网
雷火电竞官方网站
JBO竞博入口
亿博官网注册链接
壹定发游戏娱乐平台
明博电竞
JBO竞博最新官方网站
JBO官网登录入口
jbo竞博网页版登录入口
竞博体育APP官网入口
竞博jbo官网下载
竞博JBO官网在线登录
jbo竞博官方网站
竞博JBO官方在线登录
澳门买马官方网站网址
噢门彩资料网站
香港6合和彩网站入口
澳门码三一八网站
港彩图库49852b.cσm新版本历史记录查询
澳门u7彩票cc官网
澳门6合开奖结果记录
8808cc澳彩
鑫博国际app官方下载
鑫博国际彩票登入口官网
澳门鑫博国际下载
鑫博88彩票官网
鑫博国际彩票app官网下载安装
鑫博国际彩票登录入口官网
鑫广国际彩票app下载
鑫博国际彩票官网
鑫博国际彩票登录口
鑫博国际彩票APP登录
鑫彩国际彩票官网
鑫博app官方下载
大发彩票入口welcome500
博亚app官方下载
博雅彩票APP
下载博雅彩票
博雅彩票
博亚官方彩票平台手机版
博雅旧版本大全官网
博亚彩票网软件分类
博亚彩票网app介绍
博亚彩票网app使用方法
博雅彩票手机版
博亚彩票网App下载
如何在手机上直接买彩票
电子化彩票APP下载最新版本更新内容
电子化彩票app官方下载
电子化彩票app最新版本更新内容
电子化彩票app最新版本更新内容是怎样的
手机app买彩票
彩票电子app
鸿博平台网址入口官网
鸿博在线娱乐官网
鸿博官网app下载
鸿博网彩登录入口
宏博国际彩票网站下载安装
鸿博网址登录
鸿博网彩彩票注册地址
鸿博体彩app
鸿博国际app
鸿博app官网
鸿彩网app下载
鸿博app官方登录入口
绑定手机送28彩金
10块钱可以充值的棋牌
1小时赚8000元游戏
博澳体育app安卓版
博澳体育app官方网站
博冠体育app官网入口
博澳体育app
澳博在线官网娱乐游戏
博一博体育官网入口
中国足彩网官网首页查询入口
中国胜平负竞彩网
中国足球竞彩网官方发布平台
中国竞彩网官网 首页
中国体育彩票竞彩网
竞彩足彩比分500彩票网
足彩竞彩官方首页官网
比分中国足彩网
bb贝博·(中国区)有限公司官网
德信体育.(中国)官方网站
HB火博官网网址
艾弗森BB贝博平台登录入口
足球打水对冲软件
自己体彩店跟外围对刷
两个平台对打套利会被发现吗
足球刷水套利不被风控技巧
足球刷水套利交流论坛
真人体育彩票娱乐平台
澳门百乐门官网娱乐app游戏特色
庄闲游戏app官网入口
百乐门棋牌
澳门百乐门官方网站入口
真人百家乐庄闲网站
澳门棋牌在线视频观看视频
真人竞技(炸金花)
真人竞技内测平台入口
真人竞技官方正版下载
www.pgc.gov.cn
真人炸金花官方版正版下载
炸金花app游戏
10元入场的炸金花下载
真人炸金花最新版本
真人竞技炸金花下载
真人竞技炸金花挂免费下载
500体育彩票app下载
500彩票网官方下载
500足球竞彩足球比分
500彩票APP下载
500彩票app手机版下载
500彩票下载安卓app
足彩竞彩500彩票新版下载
500竞彩官网下载
彩票500app下载官网下载
C7娱乐app官方下载
C7注册入口官网注册
C7娱乐注册网站
c7游戏官网平台
C7娱乐官网
c7在线登录网址链接
C7电子娱乐官网
C7娱乐平台下载官网
C7娱乐下载官网
鄂州赌客团伙诈骗案
赌博电竞网站
官网赌博软件下载
高频彩票网送2000试玩金
赌场娱乐导航app
欧洲赌彩app下载专区
捕鱼大咖官方正版
捕鱼大咖无限破解版
超级变态捕鱼免费版
免费捕鱼
捕鱼大作战维护公告最新
捕鱼大作战官网版本
老款捕鱼游戏
经典捕鱼官方版
怀旧捕鱼下载
经典捕鱼
亚洲捕鱼官方正版app
亚洲捕鱼app官方下载最新版本更新内容
亚洲捕鱼app下载安装最新版本ios
捕鱼最新版app下载
捕鱼官方网站
亚洲捕鱼app官方最新版
乐乐捕鱼旧版本
乐乐捕鱼百度版
乐乐捕鱼app官方版
乐乐捕鱼集五福9.0
鱼乐达人微信登录
乐乐捕鱼高爆版下载
天使投资平台官网
算出彩票规律的人被赶出国家
机选有人中过一等奖吗
彩票坑人为何国家不管
中大奖的人命中注定有财运吗
彩票中奖了银行不给办卡
500彩票下载安装
竞彩足球500app下载
500万彩票官网旧版
500彩票网最新版app
旧版500彩票下载入口
500彩票app官方免费版
中国最稳十大彩票APP
赌博电子首页
体彩店微信接单怎么处罚
在微信群里买足彩可信吗
微信上购买正规体彩违法么
买体彩足球算不算赌博
微信上买体彩是赌博吗
体彩下赌注算是赌博吗
998cc娱乐彩票
959娱乐官网下载
彩票娱乐平台app下载大全最新版
赌场彩票返水最高纪录2023年最新版本更新时间
赌场彩票返水最高纪录2023年2月6日
赌场彩票返水100万判几年
赌场彩票返水标准2023年最新版本
美国彩票最高中奖纪录
中国什么彩票奖金最多
投注返水什么意思
彩票十大娱乐网站平台
凤凰娱乐彩票app下载
20年老凤凰平台登录注册
16年凤凰娱乐信誉平台
凤凰娱乐app官网入口
凤凰彩票785最新版更新内容
凤凰彩票785cc最新版3.0.0
老版凤凰785彩票APP
66导航彩票网
彩票网大全导航
彩票导航6141网址
盛世集团彩票入口导航
6617网址彩票导航彩票网3b
sss盛世集团彩票导航线路
彩票导航收录
66导航彩票网最新版
福利彩票网址导航网站查询
6617彩票导航一手
彩票站长大全导航
电竞比赛赌钱app下载
电竞赌博软件下载
有没有电竞赌博软件
线下陪玩接单平台app
C7娱乐注册入口官网注册
C7C7.ccm官方
c7官方版网站入口
南宫 c7 28 问鼎
c7·com娱乐平台下载
十大电子游戏平台网站
十大电子正规游戏网站
银河平台电子游戏
澳门pg电子游戏app
pp电子游戏平台网站
kaiyunPG电子游戏入口
pg娱乐官网
pg官方网站
麻将胡了pg
电子游戏放水规律
网赌电子游戏让你一直输
网赌电子游戏控制爆率
网赌电子游戏一直输还一直玩
电子游戏怎么赢几十万
澳门在线威尼斯9499
5845cc开元威斯尼斯人
818www威尼斯
773317电影
威尼斯886699
771771威尼斯.cm正式版
金沙乐娱场5555199vip下载
老澳门沙金娱乐app
2025开元牌棋下载
开元游戏官方网站app
开元棋下载app最新版
开元棋下载app免费
开元ky888棋app下载
999娱乐官网入口
大嘴猴棋牌dzh88cc
999棋牌入口
999pg娱乐app官网下载入口
999棋牌老版本
99棋牌最新版
999棋牌官网最新
999qp.bet棋牌
美猴王棋牌娱乐平台
赌场返水最高10万判几年
赌场返水最高处罚标准及处罚规定及标准规定详解
赌场返水50元不立案2023年最新规定解读
赌场返水一般要坐牢几年
真正赌场有反水吗
赌场返水标准
网赌一万流水返多少
在线十大娱乐棋牌平台
全国最正规的棋牌平台
bet 365网页版
best365登录网页版
bet356下载app
足球押注平台
外围买球网站入口
线上买球网址
明陞m88官网登录入口
明陞m88手机网页版
明陞m88官方网站
明陞m88备用
明陞m88备用下载地址
m88明昇体育官网备用网址
明陞M88登录入口
M88MANSION体育
明昇体育官网注册入口
明陞m88官网下载
明陞m88MANSION
龙行体育app下载官网
球探体育app官网下载
9博体育app官网入口
9博体育app
fibo飞博体育网站入口
365完美体育app官网
新澳门京葡威尼斯游戏
赌场体育返水最高纪录一览表
赌场体育高超10米返水多少米
赌场体育返水最高纪录最新消息
赌场体育最高纪录及评分标准
赌场体育最高纪录是多少
大乐透奖池最高纪录
网赌连续输是被锁定了吗
中国网赌人数有多少
赌博棋牌中国
中国赌术高手
怎么进入网赌内部
可以试玩的网赌网址
网赌模拟器在线玩
怎么才能进入网赌
20岁女网红成人平台豪赚16亿
网赌是什么
网赌视频
为什么网赌都是输
三次一个循环注码法
600永不输本金的注码技巧
一个职业赌徒的长久赌法
十赌九赢的小方法
赌场最怕三个缆法
28种注码法
一二三投注法的神奇功效
职业赌徒四大不败法则
最好的注码法赌缆111222
成功的赌徒必须遵守铁律
返水最高处罚标准2024年最新消息及处罚方法
赌场返水最高处罚标准2023年
网赌赢钱的最佳办法
赌场返水最高处罚标准2023年最新版
pg麻将胡了25000倍
真人斗地主
下载赌博软件下载
真人赌钱麻将下载安装
如何下载真人斗牛
菠菜科技
菠菜论坛有哪些
菠菜爱好者
菠菜助手
菠菜sports竞猜
新仙魔九界官方版本下载
新仙魔九界账号出售
新仙魔九界官方网站
乐乐捕鱼微信登录版本
天天仙侠捕鱼官方下载
最新捕鱼游戏.
金蟾千炮捕鱼
捕鱼技巧
捕鱼机有什么技巧打法
捕鱼小妙招
网络捕鱼技巧
菠菜彩票唯一
app送彩金58元体验金下载
菠萝彩排列三
菠萝彩排列五近1000
福建大菠萝导航官方隐藏
菠菜国际平台APP最新版本更新内容
菠菜国际app官网入口
菠菜国际app官方下载最新版本更新内容
彩投彩票app下载安装
彩虹多多官方正版下载
彩虹多多app下载彩票
彩虹多多官方官网
彩虹多多彩票app官网下载
彩虹多多
彩虹多多彩票官网下载安装
彩虹多多app官方彩票软件介绍
彩虹多多彩票免费下载
彩虹多多官方彩票下载地址
菠菜彩票网app下载最新版本更新内容
菠菜彩票网官方app最新版本介绍
菠菜网彩票
开奖总汇菠菜导航网
竞彩联盟
竞彩篮球胜负
lol联赛竞猜
lol菠菜网官网入口
英雄联盟彩票竞猜
LOL赛事竞猜
电竞赛事菠菜
菠菜娱乐游戏城5.0
三号娱乐彩票welcome
全民彩票welcome登录入口
welcome百姓彩票
在线娱乐-购彩大厅
菠菜彩票套利项目教程
网上彩票流水返点大概有多少
菠菜赚钱平台是什么
菠菜有没有正规的平台
菠菜试玩福彩3d试机号关注金码对应码
福彩3D试机号今天
菠菜3d今日试机号查询
福彩3d开机号
3D 福彩
福彩3D今天试机号码
福彩3d3D之家
菠菜3D今天开机号试机号开奖号查询
菠菜彩票APP官方版
菠菜币app官方下载最新版
菠菜彩票2024年最新版
菠菜彩票网app最新版本更新内容
菠菜彩票app官方下载
菠菜彩票app官方下载最新版
app彩票软件哪个正规
菠菜电竞app登录入口
菠菜电竞app官网入口最新版本更新内容
dota2菠菜竞猜网
企鹅电竞官方入口
vpgame还能放菠菜吗
菠菜DOTA2
u9Dota2菠菜
乐娱电竞app下载
乐娱电竞APP官网
乐娱电竞官网入口
游戏娱乐、电竞文旅
乐娱电竞官网
乐娱电竞app
缪沐阳游戏娱乐、电竞文旅
乐娱电竞下载官网
乐娱星映娱乐有限公司
电竞平台
菠菜电子平台登录入口
菠菜电子网页版入口
菠菜电子免费阅读
菠菜电子阅读器
菠菜的电子地址怎么找
菠菜电子地址怎么设置
菠菜的电子地址一般填什么
手机上的菠菜电子地址
菠菜的电子地址怎么填
菠菜的电子地址怎么添加
菠菜的电子地址怎么查找
菠菜在中国的产地
app下载安装官方免费下载
应用宝app下载
演员王东简历
沈阳市王东
王东升的个人资料
演员王东个人资料
王东是辽宁哪里人
演员王东演过什么电视剧
王东演过的电视剧
中国电子科技集团王东升
王东升在京东方的股份
沈阳重型机械厂王东简历
河南京东王东方
王东来简历
菠菜推荐平台
app体育赛事
菠菜小说网手机版
蕃茄小说TXT下载
爱下电子书 官网
电子书app下载
亚洲菠菜十大品种
亚洲菠菜品种第一名
亚洲十大菠菜排名
亚洲十大菠菜品牌
菠菜品种
菠菜最贵十大品种
菠菜十大产地排名
菠菜什么人尽量少吃
菠菜含什么元素最高
菠菜有什么营养成分
吃菠菜补什么元素
菠菜含铁量高吗
菠菜补充什么元素
菠菜含有什么维生素
菠菜补充什么营养成分
菠菜富含什么元素
菠菜的十大害处
菠菜不能和什么一起吃
菠菜的副作用和禁忌
翟老师讲菠菜从哪国传入
从外国传入中国的蔬菜
白菜什么时候传入中国
外国食物传入中国的有哪些
菠菜何时传入中国
中国本土蔬菜
胡椒什么时候传入中国
菠菜电子官网登录入口最新版本
菠菜电子app下载最新版本
菠菜电子官网APP最新版本更新内容
菠菜电子app官方下载
菠菜棋牌2024最新版本
棋牌菠菜
菠菜棋牌官方最新版本更新内容分享分享
菠菜棋牌老版本
菠菜棋牌登录入口2024
老韭菜棋牌
波克棋牌老版本v2.0
菠菜棋牌娱乐2024最新版功能介绍
菠菜棋牌娱乐改名后叫什么
菠菜娱乐2024最新版本优势评测
菠菜棋牌娱乐官网最新版使用方法
菠菜棋牌正版最新版本更新内容介绍
菠菜棋牌网页版
菠菜棋牌官网入口最新版本更新内容介绍
菠菜体育登录入口官网
菠菜体育免费登录入口
菠菜体育app官网入口最新版本更新内容
菠菜剩了多少小时不能吃了
菠菜放冰箱一个星期还能吃吗
焯过水的菠菜能放几天
菠菜在水里泡一天还能吃吗
菠菜水泡了12小时左右能吃吗
菠菜超过几小时不可以吃
吃菠菜拉肚子拉出水了
菠菜焯水后6个小时还能吃吗
菠菜淖水后时间长了有毒吗
菠菜放冰箱十天还能吃吗
菠菜真人账号注册步骤详解
菠菜真人帐号注册官网入口的最新版本更新内容
菠菜真人账号登录官网
菠菜真人账号注册的步骤及注意事项
菠菜真人账号登录2023最新版本
菠菜真人登录入口官网
菠菜真人网页版登录
菠菜真人官方网页入口链接
菠菜真人平台
菠菜真人官方版下载
菠菜真人登录app下载
菠菜真人接口的梗的背景故事与背景历史发展
菠菜真人接口的梗的来源和历史背景介绍
菠菜真人接口的梗的出处与背景探讨
菠菜真人接口的梗的出处和背景故事
api接口源码
接口和api
菠菜真人app官方入口最新版亮点
真人菠菜现金免费领取2024年最新版
真人菠菜现金免费领取2024年游戏介绍
真人菠菜现金兑换码2024最新消息
真人菠菜兑换码领取2024最新版介绍
蔬菜集团董事长简介
菠菜真人接口费用
真人菠菜手游
菠菜叔是谁
菠菜真人app最新版本更新内容与评论评价
菠菜真人国际app最新版本更新内容分享评测
菠菜真人app的最新版本更新内容详解
菠菜真人视频在线观看免费下载
菠菜真人无删减版视频
菠菜真人无码版
菠菜真人在线观看最新一期视频
菠菜真人大秀免费观看
菠菜真人无删减版
菠菜科普视频
菠菜真人成人版APP
菠菜真人免费视频播放
菠菜真人直播入口
菠菜真人在线观看官网
菠菜真人登录入口
菠菜真人官网登录入口最新版安装方法分享
香蕉水蜜桃丝瓜18岁可以吃吗
kaiyun官方网站下载手机版
云开·全站体育app登录下载
kaiyun全站网页版登录APP
开yun体育app官网入口登录入口
开yun体育app官网登录APP下载
投注返水是好还是坏
投注返水是好事吗
返水是投注流水吗
投注反水怎么算
投注返水是好兆头吗
投注反水是什么意思
投注返水最简单三个步骤
赛博体育官方网站
sportsbet赛博官网最新版功能介绍
赛博sports官方网站入口
赛博体育bet网页版入口
sportsbet官网下载
赛博体育app官网
SportyBetOnlinebetting
赛博sobets官方下载
赛博sports官方网站
赛搏sportsbet
Sportsbet官网
sportsbet赛搏体育
世博体育在线登陆入口
世博体育app下载官网
世博体育vip官网入口登录
世博国际娱乐官网网址
eaball世博登录网址
世博体育外围App下载官方版APP
世博体育下载网站
拼搏在线福彩试机号码
拼搏彩票网官网首页
彩神通官方网
拼搏在线官网彩神通系列软件
彩神通手机版官网
拼搏在线官方网站
拼博在线官方网站
拼搏在线彩神通首页
彩神通彩票软件官方网站
必一·运动(B-Sports)
B体育·(sports)
B—sports必一旧版
b体育官方下载入口最新版
bsports必一登录入口
BSport官网入口
南宫体育
博鱼·boyu体育app最新
半岛·体育官方网站
BOBSPORTS:半岛·体育(BOB)中国官方网站
半岛·体育BDSPORTS
博鱼官网app官方网站
开yunapp体育官网入口登入
博乐彩官方APP下载安装最新版本
博乐彩APP正版下载地址
博乐app下载安卓版
博乐学app
博乐app官网首页平台
老博乐棋牌平台官网
博乐填大坑官网下载app下载
168安卓版下载
355娱乐彩票官网app下载
500彩票手机版app下载
国家允许的网上彩票平台有哪些
十大购彩网站app
国家允许的购彩平台推荐
彩票平台app
国家合法购彩app最新版
十大正规彩票软件
彩虹多多app彩票官网
十大彩票平台APP
竞彩足球500完场比分
竞彩开奖最新结果
今天足彩开奖结果查询
足球赛果查询网站
竞彩足彩官网
足彩竞彩500彩票比分预测最新
足球即时比分
足球竞彩网官网平台
足彩开奖结果及奖金查询
今日竞彩的比赛结果
666cc彩票娱乐app下载
567cc彩票下载地址
380.cno玩彩网在线
玩彩网官方入口
380玩彩网官网入口网址
380玩彩网下载
380.cno玩彩网登录
千亿彩最新版本下载
9797cc彩票app
千亿彩网官方下载
千锦彩票1000亿APP下载
千娱彩票app下载
亿彩彩票平台
大发凤凰vip彩票
体彩店网上弄了个app
彩票店app下载官方
线上买彩票app平台
2025彩票app官方版
彩票实体店下单的app
线上下单彩票店出票的app
彩票软件平台app下载
JBO竞博 入口
JBO竞博·体育
JBO竞博电竞官网
JBO竞博电竞APP下载
东瀛电竞app下载
博电竞app下载苹果
极速电竞APP
博竞技app
电竞app下载地址
电子电竞app官网
中国足彩网竞彩网首页
竞彩网首页
中国竞彩彩票网
足球竞彩网官网
竞彩足球竞彩首页网
中国体育竞彩官方
中国足彩网首页官网手机版
中国竞彩彩票官网
首页 竞彩网
中国竞彩官网首页登录入口
竞彩足球胜平负计算器
(混合过关)竞彩足球计算器旧版新浪体育
中国彩票竞彩网
汇彩网app下载
重庆时时采彩app官方
汇彩网app官方网站
英皇平台app下载
pg电子娱乐平台下载安装
电子平台网站
澳门pg电子游戏网站
电子app下载官方平台
pg电子娱乐官网下载官方网站
亚洲彩票welcome登录入口
22彩票首页878.ecc
777cc彩票app最新版
亚洲彩票网址yzcp
火搏彩票是正规的吗
博艺彩彩票有风险吗
博美彩票可靠吗安全吗
美高梅旗下的博乐彩票
博亿彩票是正规平台吗
伯乐彩票怎么样
手机上怎么买彩票正规
手机购彩大厅登录入口
手机购彩-用户注册
购彩中心welcome官网登录入口
购彩中心Welcome
welcome购彩大厅用户注册
购彩大厅-购彩大厅
购彩大厅登录–官网登录
电子网站
线上电子网平台
21ic电子网
电子技术资料网站
电子线上
电子游戏在线网站
中国电子电路图网
真人线上电子平台
亚洲最大线上电子平台
电子信息网
电子科技电子网上服务大厅
澳门线上mg电子网站
牛牛官方正版免费下载
抢庄牛牛游戏app
彩·娱乐官网app
线上娱乐平台app下载
在线娱乐购彩平台
在线娱乐购彩中心
娱乐彩票官网
933彩票娱乐平台特色
好搏体育app官方下载
好博体育app在线下载
好博app官方下载
好博体育app官网
赛酷体育下载官网
体育app
hg8088皇冠新体育
https://ag.hg050.com
welcome皇冠地址登录入口
皇冠app正版下载入口
HG08皇冠登录入口
2串1稳赢套利妙法
二串一一年稳赢100万
两个平台对打套利技巧
负盈利套利7个月3000多万
竞彩二串一对冲套利
YOBO官方网站
AG 真人百家乐注册
8868体育官方网站app
娱乐彩票app官网下载
娱乐彩票全平台
真人娱乐入口注册平台有哪些
彩票平台第一娱乐
天天娱乐彩票注册
在线娱乐彩票注册平台
多宝彩票娱乐平台
澳客app官方正版下载
澳博官方网站APP
澳博app官方入口
澳客网app
澳博集团彩票平台官网
澳博app安卓下载
澳博手机版官网
澳博app官方下载安装
澳博集团app下载
竞彩足球混合投注
足彩竞彩计算器
竞彩足球混合投注500
竞彩足球胜负彩投注
竞彩足球计算器app
竞彩足球赛程
计算器竞彩足球胜平负
游戏赚钱
赚钱的小游戏
棋牌软件制作开发
网赌怎么举报有效
赌博的网站登不了
应用试客app下载
钱咖试玩app下载
试客小兵官网入口
钱行app试玩软件
手机试玩平台app
试玩app赚钱平台排行榜
皇家捕鱼金猴爷下载
捕鱼充值折扣平台
皇家捕鱼电玩城有安卓版吗
皇家电玩城手机版官网
金猴爷捕鱼电玩城
金猴爷皇家电玩官网版
hydwcom电玩城官网版
官方正版电玩城捕鱼
柳岩代言捕鱼大作战官网下载
捕鱼大玩咖官方正版
0.01折手游官方网站
捕鱼能上下分10元起下
捕鱼微信秒到账2024
捕鱼大世界官方正版下载
捕鱼大世界官方网站
捕鱼世界唯一官方下载
捕鱼大世界官网下载
捕鱼大世界官方2025最新版
下载捕鱼正版手游
捕鱼大世界最新版本
捕鱼大世界官方正版
捕鱼大世界
网络捕鱼游戏
电玩捕鱼app
人气最火的捕鱼手游
2025最火捕鱼手游推荐
2025好玩的捕鱼手游
最新捕鱼游戏推荐
正版休闲捕鱼手游
澳门赌鱼地址一览表2023最新版
澳门赌鱼一览表
2024年最新赌鱼地址
澳门赌鱼地址大全2023最新版
查钓鱼网站
打鱼赢钱的网站
投资20元一小时赚500
捕鱼充值游戏平台
捕鱼游戏app
捕鱼炸翻天
捕鱼输钱就上头,一直想充值
网赌捕鱼输了200万赢回来了
打鱼游戏打鱼平台赚钱
能提现的打鱼游戏 正版的
下载打鱼可以提现
天天真人捕鱼
途游捕鱼柳岩
爆率超高的捕鱼游戏
最容易爆金币的捕鱼游戏
捕鱼手游哪款爆率最高
良心捕鱼游戏推荐
2025最新捕鱼游戏
腾讯捕鱼手游
王者捕鱼
趣游捕鱼官网
全部的捕鱼游戏大全
56677凤凰平台旧版本
959娱乐3.0最新版本更新内容
cp55彩票入口
2025彩票app下载v.9.9.9版
PG网赌app
6566cc天天彩天彩网
赌彩app
cc天天彩网官网网址
激情综合赌
赌彩平台下载
综合平台彩票
手机赌彩的app
网赌每天赢100
正版106cc彩票app下载
106cc彩票app旧版安装
彩票106 cc手机安卓版
106老版彩票官网v1.0
106cc官网彩票下载
106娱乐彩票app软件下载4.0
106cc彩票1.0.2
106旧版本彩票娱乐
106娱乐彩票的官方网站
106官网彩票安卓版2.0
106官网彩票下载安卓版
106官网彩票下载
竞彩足球胜平负计算器新浪
今天竞彩足球胜平负比赛结果310
竞彩混合过关和胜平负
计算器足球计球胜平负游戏
(混合过关)竞彩足球计算器
今天足彩竞彩
中国足球彩票
足彩 胜负彩
800cc彩票网登录
800CC彩票登录入口官网
800cc彩票官网下载入口
800cc彩票平台最新版本更新内容
800彩票平台app下载
彩票网址:800.cc/800.show
800cc彩票网站
800cc彩票3.0大厅
8000cp666cc
800万彩票平台入口
电竞娱乐app下载安装
电竞娱乐app下载官网最新版安装方法
电竞娱乐官网入口登录注册流程详解
电子娱乐平台app下载
电竞娱乐app下载安装最新版2023
泛娱乐电竞
电竞体育
澳门网站电子游戏大全
lol比赛怎么买竞彩
lol比赛押注
lol比赛买输赢的网站
十大电竞竞猜app
导航到南京电竞馆
网鱼电竟是什么
网鱼网咖价格表2025
网鱼电竞酒店官网
网鱼电竞网鱼网咖
网鱼官方网站
网鱼网咖app下载
导航到网鱼电竞酒店
网鱼电竞官网
网鱼app下载
网鱼电竞酒店
网鱼网咖app
下载送38元彩金有哪些
注册就送68无需首充的平台
送28彩金平台
澳门pg电子注册送38彩金
威尼斯注册送38元
5期必中的倍投法
121212倍投最简单方法
500元倍投16期方案
每天赢300赢了一年
红蓝最厉害三个打法
最聪明的倍投方法
电子pg电子模拟器网站
麻将胡了网页版入口
PG电子麻将胡了2
赌博电子登录
赌博电子集团
赌博电子最新
贵州省较大数额罚款 标准
河南卡车堵洪水赔偿
偷倒污水140吨判刑吗
河南水灾赔偿标准 2021
2021年河南水灾国家补偿标准
2021洪水受灾国家补偿标准
闯红灯最新处罚规定
2023年交通罚款总额
2019年醉驾新规定怎样处罚的
2020长江全面禁渔补偿政策
司机错过高速出口倒车被罚
云南醉驾处罚最新消息
打牌赚现金
打牌赚钱游戏可提现
打牌提现的软件
利博体育官网入口网址
利博体育官网登录
利博官网登录入口
利博体育在线平台
利博百家乐
利博登录
利博官方网站
博体育平台app下载
南宫·NG体育
ng体育注册平台入口
ng电子游戏官网入口
NG体育app下载入口
南宫NG28平台登录入口
www.ngty01.cn
NG28.666官网版
ng体育娱乐app下载
ng28.c7.28
南宫·NG28下载链接
NG体育官网链接入口
NG体育app下载安装
十大信誉赌博官网
赌博体育首页
NG体育娱乐网官方下载
九游娱乐app官方入口
BB体育登录入口app
亚盈体育app下载入口
体育娱乐平台app大全
ku.bet9.com
十大体育娱乐网址
ayx体育官网入口链接
赌博体育综合
博体育app下载官网
OB·体育app下载
冠军体育app官方网站入口
赌场体育返水最新消息今天
赌场体育最新规定2024年最新消息
2024年赌场体育返水新规定及处罚详解详解
2024年赌场体育新规定
2023年抓赌新规定
2025年抓赌新规定
2020年麻将抓赌新规定
www.55123.com.澳精准资料查询
澳门特马最准网站
澳门一马一肖
澳门六会彩,今晚开什么号
澳门网站一码期期中
49.ccm澳门开奖
澳中特网com
49.ccm澳彩资料图库
澳门特马开奖结果
刑事拘留37天后不放人怎么办
开设赌场要几人口供才能定罪
开设赌场不是主犯,最多能判多少
取保后最怕三个征兆
开设赌场罪的定罪标准
2025最新开设赌场量刑
开设赌场罪16万退了能缓刑吗
真人四人麻将免费版
真人麻将app下载
欢乐真人麻将免费下载
麻将真人麻将微信付款
欢乐真人麻将
真人麻将app哪个好
麻将游戏真人美女
赌博体育唯一
真人扎金花现金
真人百家乐注册
真人AG百家乐
赌博体育试玩
十大返水平台排名
返水最厉害三个平台
返水最严重的三个平台
返水最厉害的三个app
返水最怕三个平台
返点高的娱乐平台
赌博棋牌在线
AG真人百家乐官方网站
sport体育平台
菠菜sports官网入口
菠菜sports安卓版下载
菠菜sports官网中文版下载安装
菠菜sports官网最新版下载
菠菜sports最新版下载链接
菠菜sports网页版登录入口
菠菜sports官网中文版的功能介绍
菠菜sportsapp最新版本更新内容分享方法
菠菜sports免费版下载
电玩捕鱼3d版
电玩捕鱼手机版安卓版
电玩捕鱼官网
菠菜捕鱼app下载安装最新版本更新内容
菠菜捕鱼app最新版本更新内容是什么
菠菜捕鱼最新版本下载安装方法
亚洲菠菜app官方下载
亚洲菠菜app官方入口
亚洲菠菜app安装
亚洲菠菜APP入口
亚洲菠菜网app下载安装最新版本优势
亚洲菠菜app最新版特色功能介绍
菠菜返水最忌三个原因
菠菜捕鱼返水最简单三个步骤
菠菜捕鱼返水规律口诀与技巧详解
菠菜返水最简单的三个步骤
菠菜捕鱼返水最简单三个技巧
捕鱼大咖攻略
捕鱼手游直播间
捕鱼大作战直播间
捕鱼大咖官方直播
捕鱼大咖直播间现场直播
乐乐捕鱼直播间
鱼乐达人
乐乐捕鱼直播在线观看
鱼乐达人主播顿顿
菠菜彩票地址
菠菜集团
888菠菜集团
菠菜彩票正版APP
菠菜福彩3D下载
菠菜彩票APP最新版更新内容
菠菜彩票官网最新版下载
polocai中国官方网站
http://bbs.polocai.com/thread
polocai官网中文版
www.polocai.com
//bbs.polocai.com
polocai网页版登陆
bbs.polocai.com
polocai中国官网中国官方下载
菠菜彩票官方app下载
菠菜彩票网app官方最新版本功能介绍与功能展示
菠菜彩票网app官方下载安装最新版本2023年
菠菜彩票网app下载安装
彩票网官方app下载
(混合过关)竞彩网计算器旧版新浪成都
竞彩足球500彩票网
500足彩比分完整比分
(混合过关)竞彩网计算器足球旧版
500足彩比分网
彩票民间高手的方法
中国体育彩票网官方网站
中国彩票查询结果
中国体育彩票官方首页
中国彩票网站官网入口
中国福彩彩票官网
中国彩票
中国福彩网官网首页
中国彩票网下载
中国福彩官方首页
菠菜试玩app下载安装最新版本更新内容
菠菜试玩app改名了吗
菠菜试玩app最新版本更新内容分享
绿菠菜账号如何下载
菠菜是什么平台
菠菜城市
菠菜管理系统软件语言
菠菜竞彩app官方
菠菜彩票app免费进入
在线娱乐彩票大厅～welcome
菠菜彩票APP最新版本更新内容解读
菠菜电竞app最新版本更新内容如何查看
菠菜电竞app免费下载安装最新版本更新内容汇总
菠菜电竞网页版入口2023最新版亮点
菠菜电竞网页版入口官网版特色
菠菜电竞网页版入口最新版本更新内容
雷速体育app官方下载
雷火官网app下载
雷火电竞lh登录入口
雷火官网入口
雷火体育APP登录入口
LH雷火官网
雷火官网
菠菜电竞主页
菠菜电竞主页最新消息
菠菜电竞最新视频2023年
score电竞数据官网
菠菜电竞改名后叫啥名字2023年最新消息
菠菜电竞改名后叫啥软件2023最新版使用方法
菠菜电竞改名最新消息
菠菜商城登录入口
菠菜电子商城app下载
菠菜电子商城app官方下载安装
菠菜电子商城app官方下载安卓
黑菠菜网
菠菜电子版
国家电子教材网(官网)
菠菜电子书app最新版本更新内容
菠菜电子书免费阅读全文
菠菜电子书app官方下载最新版本更新内容
菠菜电子娱乐官网入口
菠菜娱乐app官网
菠菜电子娱乐app最新版本更新内容
菠菜电子娱乐app下载官方正版
99499www威尼斯游戏背景
亚洲必赢登录游戏平台
亚洲十大搏彩公司排名
yeezy亚洲2022版
亚洲必赢bmw网址
菠菜含铁高的错误由来
菠菜营养成分表100g
菠菜含铁量
菠菜的营养成分
菠菜是哪个国家传入中国的
电竞菠菜下载
线上菠菜下载
菠菜棋牌官网入口最新版本更新内容
菠菜棋牌官方入口地址
菠菜棋牌官方网页
菠菜棋牌改名后叫什么名字2024年
菠菜棋牌改名后叫什么名字2023年最新版本更新时间
菠菜棋牌2024年最新版
大唐棋牌改名了最新消息
盛宴棋牌改名后叫什么
最新菠菜优惠论坛
菠菜地址大全
菠菜拉新
新兰花菠菜
菠菜影视的最新链接
菠菜的新吃法
野菠菜的最新信息
棋牌网站大全
空间棋牌2020旧版
蓝洞棋牌6.7.3(711)版本
一诺棋牌2024最新版本
神赚棋牌最新6.1.0版
空间棋牌6.1.0版本
13水棋牌平台
大放水棋牌官网入口
2025放水口子
大好玩棋牌560cc8888最新版特色服务
新平台大放水娱乐网站
大放水彩票app
棋牌软件 平台
齐齐乐棋牌官方正版
齐齐乐手游平台app
齐齐乐老版本下载
齐齐乐棋牌最新官网
齐齐乐官网app
齐齐乐棋牌ios版官网
齐齐乐炸金花官网
老版齐齐乐炸金花
齐齐乐炸金花ios版下载
棋乐棋牌旧版本
菠菜游戏正版
菠菜宝官网
菠菜娱乐app官方入口最新版本更新内容
菠菜娱乐在线观看
菠菜娱乐app官方下载
菠菜娱乐APP
菠菜体育app官方正版
365wm完美体育官网app下载
完美体育365wm最新版本更新内容
完美体育在线登录入口
完美体育网页版入口链接
完美体育官网
完美体育app官方下载
完美(中国)体育官方网站
wm365wmvip登录平台
wm完美体育在线登录
完美体育网页版登录
365wn完美体育
菠菜真人集团董事长简介
菠菜真人集团董事长是谁
菠菜真人集团董事长女儿
菠菜真人集团三个创始人介绍
菠菜真人集团是正规公司吗
菠菜真人集团张健简介
菠菜真人创始人背景资料与成就
线上真人安卓版app下载
手机真人app安卓版下载
国际真人手机官网app下载
菠菜真人视频高清版在线播放
菠菜真人在线播放
菠菜真人视频免费播放下载安装
菠菜视频大全免费观看
菠菜视频免费观看高清下载安装
菠菜真人竞猜app最新版本更新内容
菠菜真人竞猜APP最新版本更新内容分享
菠菜真人竞猜app最新版下载安装方法
菠菜真人娱乐有套路吗
菠菜真人娱乐跑路了
菠菜真人娱乐倒闭了吗最新消息
菠菜真人娱乐倒闭了吗知乎
菠菜真人娱乐是正规公司吗知乎
菠菜真人娱乐是杂牌吗
娱乐系统平台开发菠菜
菠菜投注最准确的软件
真人菠菜投注app最新版本更新内容
菠菜投注app官方下载最新版本更新内容介绍
菠菜真人视频无删减版下载
菠菜真人动漫在线播放免费漫画
菠菜真人动漫无删减版观看
菠菜科普视频动画
大力水手菠菜动画片
菠菜真人动漫在线播放
'''
keywords = [keyword.strip() for keyword in KEYWORDS.split('\n') if keyword.strip()]
domains = """
http://www.hmc.edu.cn/
http://www.zjhu.edu.cn/
http://www.wzut.edu.cn/
http://www.zjcm.edu.cn/
http://www.zjpu.edu.cn/
http://www.jhc.edu.cn/
"""
domain_list = [domain.strip() for domain in domains.split('\n') if domain.strip()]

# 配置日志
logging.basicConfig(filename='scraper.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', encoding='utf-8')

# 提取元数据
def extract_metadata(soup):
    title = soup.find('title').text.strip() if soup.find('title') else ''
    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
    keywords_content = keywords_meta['content'].strip() if keywords_meta else ''
    description_meta = soup.find('meta', attrs={'name': 'description'})
    description_content = description_meta['content'].strip() if description_meta else ''
    return {'title': title, 'keywords': keywords_content, 'description': description_content}

# 提取“大家还在搜”数据
def extract_related_searches(soup):
    related_div = soup.find('div', class_='c-title', string='大家还在搜')
    if related_div and related_div.next_sibling:
        data_tool = related_div.next_sibling.find('div', class_='sc-feedback')
        if data_tool and 'data-tool' in data_tool.attrs:
            try:
                tool_data = json.loads(data_tool['data-tool'].replace("'", '"'))
                related_words = tool_data.get('feedback', {}).get('suggest', {}).get('ext', {}).get('relation_words', '')
                return related_words
            except json.JSONDecodeError:
                return ''
    return ''

# 检测内容类型
def detect_content_type(url, soup):
    if "article" in url.lower() or soup.find('article'): return "文章"
    elif "product" in url.lower() or soup.find('div', class_=re.compile('product')): return "产品页面"
    elif "forum" in url.lower() or soup.find('div', class_=re.compile('post|thread')): return "论坛"
    elif "video" in url.lower() or soup.find('video'): return "视频"
    else: return "未知"

# 检查是否为响应式网站
def check_responsive(url):
    try:
        soup = BeautifulSoup(requests.get(url, timeout=5).text, 'html.parser')
        return "响应式网站" if soup.find('meta', attrs={'name': 'viewport'}) else "非响应式网站"
    except Exception as e:
        return f"无法访问: {e}"

# 获取域名最后更新时间
def get_domain_last_updated(url):
    try:
        domain = url.split("//")[-1].split("/")[0]
        updated_date = whois.whois(domain).updated_date
        if isinstance(updated_date, list):
            updated_date = updated_date[0]
        return updated_date.replace(tzinfo=None) if updated_date and hasattr(updated_date, 'tzinfo') else updated_date
    except Exception:
        return None

# 获取品牌词
def get_brand_terms(website_name):
    try:
        response = requests.get(website_name, timeout=5)
        response.encoding = response.apparent_encoding
        title = BeautifulSoup(response.text, 'html.parser').find('title')
        if title and title.text.strip():
            return ', '.join(filter(None, re.split(r'[\s\W_]+', title.text.strip())))
        domain = website_name.split("//")[-1].split("/")[0].replace('www.', '').split('.')[0]
        return domain or "未识别品牌词"
    except Exception:
        return ""

# 搜索关键词
async def search_keyword(page, keyword, all_results):
    search_url = f"https://m.baidu.com/s?wd={keyword}"
    try:
        await page.goto(search_url, timeout=30000)
        await page.wait_for_selector("div.result", timeout=30000)
        html = await page.content()
        soup = BeautifulSoup(html, 'html.parser')
        related_searches = extract_related_searches(soup)
        results = soup.select('div.result')[:2]
        for result in results:
            data_log = result.get('data-log', '{}')
            link = eval(data_log.replace("'", '"')).get('mu', '无链接') if data_log else '无链接'
            title_element = result.select_one('h3') or result.select_one('a[class*="title"]')
            title_text = title_element.text.strip() if title_element else ''
            metadata, content_type = ({'title': '', 'keywords': '', 'description': ''}, '未知') if link == '无链接' else await fetch_page_data(page, link)
            # 添加UCI计算
            uci = await calculate_page_uci(page, link)
            # 添加新功能调用
            responsive_type = check_responsive(link) if link != '无链接' else "无链接"
            last_updated = get_domain_last_updated(link) if link != '无链接' else None
            brand_terms = get_brand_terms(link) if link != '无链接' else ""
            all_results.append({
                '搜索关键词': keyword,
                '大家还在搜': related_searches,
                '内容类型': content_type,
                '网址': link,
                '页面标题': metadata['title'],
                '关键词': metadata['keywords'],
                '描述': metadata['description'],
                'UCI': uci,
                '响应类型': responsive_type,
                '最后更新时间': last_updated,
                '品牌词': brand_terms
            })
    except Exception as e:
        logging.error(f"关键词 {keyword} 搜索失败：{e}")
    print(f"完成关键词：{keyword}")

# 获取页面数据
async def fetch_page_data(page, link):
    try:
        await page.goto(link, timeout=10000)
        await page.wait_for_selector("body", timeout=10000)
        page_content = await page.content()
        page_soup = BeautifulSoup(page_content, 'html.parser')
        return extract_metadata(page_soup), detect_content_type(page.url, page_soup)
    except Exception as e:
        logging.error(f"链接 {link} 访问失败：{e}")
        return {'title': '', 'keywords': '', 'description': ''}, '未知'

# 获取域名数据
async def get_domain_data(page, domain):
    try:
        await page.goto(domain, timeout=10000)
        await page.wait_for_selector("body", timeout=10000)
        content = len(await page.content())
        links = len(await page.query_selector_all("a"))
        return content, links
    except Exception as e:
        logging.error(f"域名 {domain} 访问失败：{e}")
        return 0, 0

# 计算 UCI
def calculate_uci(cf, lf, df, uf):
    return min(100, round(0.4 * cf + 0.3 * lf + 0.2 * df + 0.1 * uf, 2))

# 计算页面UCI
async def calculate_page_uci(page, link):
    try:
        content, links = await get_domain_data(page, link)
        cf = min(100, content / 10000 * 100)
        lf = min(100, links / 50 * 100)
        df = 100 if tldextract.extract(link).registered_domain in ["baidu.com", "qq.com", "163.com", "sohu.com", "sina.com"] else 0
        uf = 50
        uci = calculate_uci(cf, lf, df, uf)
        return uci
    except Exception as e:
        logging.error(f"计算页面 {link} 的UCI失败：{e}")
        return 0

# 处理域名
async def process_domain(page, domain):
    content, links = await get_domain_data(page, domain)
    cf = min(100, content / 10000 * 100)
    lf = min(100, links / 50 * 100)
    df = 100 if tldextract.extract(domain).registered_domain in ["baidu.com", "qq.com", "163.com", "sohu.com", "sina.com"] else 0
    uf = 50
    uci = calculate_uci(cf, lf, df, uf)
    return {"域名": domain, "UCI": uci}

# 主函数
async def main():
    start_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行开始时间 {start_time}")
    logging.info(f"运行开始时间 {start_time}")

    all_results = []
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_file = f"{script_name}_{time.strftime('%Y-%m-%d_%H.%M')}.xlsx"

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False)
        page = await browser.new_page()

        # 处理关键词
        for i, keyword in enumerate(keywords, 1):
            await search_keyword(page, keyword, all_results)
            if i % 5 == 0 or i == len(keywords):
                df = pd.DataFrame(all_results)
                # 清理非法字符
                df = df.replace(ILLEGAL_CHARACTERS_RE, "", regex=True)
                print(f"\n当前数据结果（前5条）：\n{df.tail(5).to_string()}")
                print(f"完成进度：{i}/{len(keywords)} 个关键词")
                if i == 5:
                    df.to_excel(excel_file, index=False, sheet_name='Sheet1', engine='openpyxl')
                else:
                    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df.to_excel(writer, index=False, sheet_name='Sheet1')
                print(f"结果已保存至 {excel_file} 的 Sheet1")

        # 处理域名
        domain_results = [await process_domain(page, domain) for domain in domain_list]
        df_domains = pd.DataFrame(domain_results)
        # 清理非法字符
        df_domains = df_domains.replace(ILLEGAL_CHARACTERS_RE, "", regex=True)
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_domains.to_excel(writer, index=False, sheet_name='Sheet2')
        print(f"域名数据已保存至 {excel_file} 的 Sheet2")

        await browser.close()

    # 格式化 Excel
    wb = load_workbook(excel_file)
    ws1, ws2 = wb['Sheet1'], wb['Sheet2']
    ws1.freeze_panes, ws2.freeze_panes = 'A2', 'A2'
    ws1.auto_filter.ref, ws2.auto_filter.ref = ws1.dimensions, ws2.dimensions
    desc_col_idx = list(df.columns).index('描述') + 1
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=desc_col_idx, max_col=desc_col_idx):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    # 将 Sheet1 的“大家还在搜”列复制到 Sheet3，按 & 分列，合并成单列并去重
    ws3 = wb.create_sheet('Sheet3')
    related_col_idx = list(df.columns).index('大家还在搜') + 1  # 找到“大家还在搜”列的索引
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=related_col_idx, max_col=related_col_idx):
        for cell in row:
            ws3[f'A{cell.row - 1}'].value = cell.value  # 复制到 Sheet3 的 A 列

    # 使用 split 分列并合并成单列
    all_values = []
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value and '&' in str(cell.value):
                split_values = str(cell.value).split('&')
                all_values.extend(split_values)
            elif cell.value:
                all_values.append(str(cell.value))

    # 去重并写入 Sheet3 的 A 列
    unique_values = list(dict.fromkeys(all_values))  # 去重，保留顺序
    ws3.delete_rows(1, ws3.max_row)  # 清空 Sheet3
    for i, value in enumerate(unique_values, 1):
        ws3[f'A{i}'].value = value

    wb.save(excel_file)
    print(f"Excel 文件已格式化，且 Sheet1 的‘大家还在搜’列已复制到 Sheet3，按 & 分列，合并并去重")

    end_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行结束时间 {end_time}")
    logging.info(f"运行结束时间 {end_time}")

if __name__ == "__main__":
    asyncio.run(main())
