import asyncio
import logging
import re
import time
import pandas as pd
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook
import os

# 关键词列表
KEYWORDS = '''
918博天堂btt
918博天堂官方app
'''
keywords = [keyword.strip() for keyword in KEYWORDS.split('\n') if keyword.strip()]

# 配置日志
logging.basicConfig(
    filename='scraper.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)

# 辅助函数：提取元数据
def extract_metadata(soup):
    title = soup.find('title').text.strip() if soup.find('title') else ''
    keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
    keywords_content = keywords_meta['content'].strip() if keywords_meta else ''
    description_meta = soup.find('meta', attrs={'name': 'description'})
    description_content = description_meta['content'].strip() if description_meta else ''
    return {'title': title, 'keywords': keywords_content, 'description': description_content}

# 辅助函数：检测内容类型
def detect_content_type(url, soup):
    if "article" in url.lower() or soup.find('article'):
        return "文章"
    elif "product" in url.lower() or soup.find('div', class_=re.compile('product')):
        return "产品页面"
    elif "forum" in url.lower() or soup.find('div', class_=re.compile('post|thread')):
        return "论坛"
    elif "video" in url.lower() or soup.find('video'):
        return "视频"
    else:
        return "未知"

# 搜索关键词并提取数据
async def search_keyword(page, keyword, all_results):
    search_url = f"https://m.baidu.com/s?wd={keyword}"

    try:
        await page.goto(search_url, timeout=30000)
        await page.wait_for_selector("div.result", timeout=30000)
        html = await page.content()
        soup = BeautifulSoup(html, 'html.parser')
        results = soup.select('div.result')[:5]  # 只取前5个结果

        for result in results:
            data_log = result.get('data-log', '{}')
            try:
                link = eval(data_log.replace("'", '"')).get('mu', '无链接')
            except (SyntaxError, NameError, TypeError):
                link = '无链接'
            title_element = result.select_one('h3') or result.select_one('a[class*="title"]')
            title_text = title_element.text.strip() if title_element else ''

            if link != '无链接':
                try:
                    await page.goto(link, timeout=10000)
                    await page.wait_for_selector("body", timeout=10000)
                    page_content = await page.content()
                    page_soup = BeautifulSoup(page_content, 'html.parser')
                    metadata = extract_metadata(page_soup)
                    content_type = detect_content_type(page.url, page_soup)
                except Exception as e:
                    logging.error(f"关键词 {keyword}，链接 {link} 访问失败：{e}")
                    metadata = {'title': '', 'keywords': '', 'description': ''}
                    content_type = '未知'
            else:
                metadata = {'title': '', 'keywords': '', 'description': ''}
                content_type = '未知'

            all_results.append({
                '搜索关键词': keyword,
                '内容类型': content_type,
                '网址': link,
                '页面标题': metadata['title'],
                '关键词': metadata['keywords'],
                '描述': metadata['description'],
            })
    except Exception as e:
        logging.error(f"关键词 {keyword} 搜索失败：{e}")
    print(f"完成关键词：{keyword}")

# 主函数
async def main():
    start_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行开始时间 {start_time}")
    logging.info(f"运行开始时间 {start_time}")

    all_results = []
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    excel_file = f"{script_name}_{time.strftime('%Y-%m-%d_%H.%M')}.xlsx"

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(headless=False)
        page = await browser.new_page()

        for i, keyword in enumerate(keywords):
            await search_keyword(page, keyword, all_results)

            if (i + 1) % 5 == 0 or (i + 1) == len(keywords):
                df = pd.DataFrame(all_results)
                if not os.path.exists(excel_file):  # 检查文件是否存在
                    df.to_excel(excel_file, index=False, engine='openpyxl')
                else:
                    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

                print(f"已处理 {i + 1}/{len(keywords)} 个关键词，结果已保存至 {excel_file}")

        await browser.close()

    # 加载 Excel 文件并应用格式
    wb = load_workbook(excel_file)
    ws = wb.active

    # 冻结首行和设置筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 设置“描述”列右对齐
    desc_col_idx = list(df.columns).index('描述') + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=desc_col_idx, max_col=desc_col_idx):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    # 保存格式化后的 Excel 文件
    wb.save(excel_file)
    print(f"Excel 文件已格式化")

    end_time = time.strftime("%Y-%m-%d %H:%M")
    print(f"运行结束时间 {end_time}")
    logging.info(f"运行结束时间 {end_time}")

if __name__ == "__main__":
    asyncio.run(main())