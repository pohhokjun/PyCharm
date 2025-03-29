import logging
import time
import random
import json
import re
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
from selenium.common.exceptions import WebDriverException, TimeoutException, NoSuchWindowException
import undetected_chromedriver as uc
import pickle
import os
import pandas as pd
from urllib.parse import urlparse
from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook

# 关键词列表
KEYWORDS = '''
918博天堂btt
918博天堂官方app
'''
keywords = [keyword.strip() for keyword in KEYWORDS.split('\n') if keyword.strip()]

# 配置日志
logging.basicConfig(
    filename='scraper.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)


# 设置浏览器驱动
def setup_driver(disable_js=False):
    chrome_options = uc.ChromeOptions()
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--disable-infobars")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--log-level=3")
    chrome_options.add_argument(
        "--user-agent=Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 BaiduHD/6.11.0.11"
    )
    if disable_js:
        chrome_options.add_experimental_option("prefs", {"profile.managed_default_content_settings.javascript": 2})

    driver = uc.Chrome(options=chrome_options, headless=False)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
        """
    })
    return driver


# 加载cookies
def load_cookies(driver, cookies_file):
    if os.path.exists(cookies_file):
        driver.get("https://m.baidu.com")
        with open(cookies_file, "rb") as f:
            cookies = pickle.load(f)
            for cookie in cookies:
                driver.add_cookie(cookie)
        driver.refresh()


# 保存cookies
def save_cookies(driver, cookies_file):
    with open(cookies_file, "wb") as f:
        pickle.dump(driver.get_cookies(), f)


# 提取真实URL
def extract_real_url(driver, link):
    if "baidu.com/link?url=" in link:
        driver.execute_script("window.open(arguments[0]);", link)
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(3)
        real_url = driver.current_url
        driver.close()
        driver.switch_to.window(driver.window_handles[0])
        return real_url
    return link


# 判断网站内容类型
def detect_content_type(url, html):
    if "article" in url.lower() or soup.find('article'):
        return "文章"
    elif "product" in url.lower() or soup.find('div', class_=re.compile('product')):
        return "产品页面"
    elif "forum" in url.lower() or soup.find('div', class_=re.compile('post|thread')):
        return "论坛"
    elif "video" in url.lower() or soup.find('video'):
        return "视频"
    else:
        return "未知"


# 判断两个URL是否实质相同
def is_same_content_url(url1, url2):
    parsed_url1 = urlparse(url1)
    parsed_url2 = urlparse(url2)
    core_domain1 = parsed_url1.netloc.replace('www.', '')
    core_domain2 = parsed_url2.netloc.replace('www.', '')
    core_url1 = core_domain1 + parsed_url1.path
    core_url2 = core_domain2 + parsed_url2.path
    if core_url1 == core_url2:
        return True
    path_parts1 = parsed_url1.path.split('/')
    path_parts2 = parsed_url2.path.split('/')
    common_parts = set(path_parts1) & set(path_parts2)
    return len(common_parts) > 1 and any(part for part in common_parts if len(part) > 5)


# 初始化浏览器
driver_no_js = None
driver_with_js = None
try:
    logging.info("初始化浏览器（禁用 JavaScript）...")
    driver_no_js = setup_driver(disable_js=True)
    logging.info("初始化浏览器（启用 JavaScript）...")
    driver_with_js = setup_driver(disable_js=False)
    cookies_file = "baidu_cookies.pkl"
    load_cookies(driver_with_js, cookies_file)

    excel_file = "关键词.xlsx"
    first_write = True  # 标记是否是第一次写入

    for keyword in keywords:
        all_excel_data = []  # 每次关键词清空列表
        search_url = f"https://m.baidu.com/s?wd={keyword}"
        print(f"\n=== 搜索关键词: {keyword} ===")
        logging.info(f"搜索关键词: {keyword}")

        print("页面加载中...")
        max_retries = 3
        for attempt in range(max_retries):
            try:
                driver_with_js.get(search_url)
                ActionChains(driver_with_js).move_by_offset(random.randint(50, 100), random.randint(50, 100)).perform()
                WebDriverWait(driver_with_js, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.result")))
                break
            except TimeoutException:
                print(f"页面加载超时（第 {attempt + 1}/{max_retries} 次尝试），可能触发验证码或网络问题！")
                logging.error(f"页面加载超时: {search_url} （第 {attempt + 1} 次尝试）")
                if attempt == max_retries - 1:
                    print("达到最大重试次数，跳过此关键词...")
                    logging.error(f"达到最大重试次数，跳过关键词: {keyword}")
                    break
                time.sleep(random.uniform(3, 5))
            except Exception as e:
                print(f"加载页面失败: {e}")
                logging.error(f"加载页面失败: {e}")
                break

        if "div.result" not in driver_with_js.page_source:
            continue

        save_cookies(driver_with_js, cookies_file)

        print("滚动页面以加载动态内容...")
        driver_with_js.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(4, 6))

        current_url = driver_with_js.current_url
        page_title = driver_with_js.title
        print(f"当前URL: {current_url}")
        print(f"页面标题: {page_title}")
        logging.info(f"当前URL: {current_url}, 页面标题: {page_title}")

        max_attempts = 3
        attempt = 0
        while (
                "captcha" in current_url.lower() or "404" in page_title.lower() or "not found" in page_title.lower()) and attempt < max_attempts:
            print("检测到验证码或404错误！")
            logging.warning("检测到验证码或404错误")
            print("请手动完成验证码验证（如果有），然后按Enter键继续，或关闭浏览器重试...")
            input()
            print(f"第 {attempt + 1} 次重新加载搜索页面...")
            driver_with_js.get(search_url)
            WebDriverWait(driver_with_js, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.result")))
            driver_with_js.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(4, 6))
            current_url = driver_with_js.current_url
            page_title = driver_with_js.title
            print(f"当前URL: {current_url}")
            print(f"页面标题: {page_title}")
            logging.info(f"重新加载后 - 当前URL: {current_url}, 页面标题: {page_title}")
            attempt += 1

        if "captcha" in current_url.lower() or "404" in page_title.lower() or "not found" in page_title.lower():
            print("多次尝试后仍未加载正确页面，请检查网络或使用代理！")
            logging.error(f"多次尝试后仍未加载正确页面，跳过关键词: {keyword}")
            continue

        html = driver_with_js.page_source
        soup = BeautifulSoup(html, 'html.parser')
        results = soup.select('div.result')

        search_results = []
        if not results:
            print("未找到搜索结果！可能原因：页面结构变化、验证码或网络问题。")
            logging.warning("未找到搜索结果")
        else:
            for i, result in enumerate(results[:7], 1):  # 只取前7个结果
                data_log = result.get('data-log', '{}')
                try:
                    data_log_json = json.loads(data_log.replace("'", '"'))
                    link = data_log_json.get('mu', '无链接')
                except json.JSONDecodeError:
                    link = '无链接'

                title_element = result.select_one('h3') or result.select_one('a[class*="title"]')
                title_text = title_element.text.strip() if title_element else ''

                real_link = extract_real_url(driver_with_js, link)
                logging.info(f"结果 {i}: 标题: {title_text}, 链接: {real_link}")
                search_results.append({
                    'original_keyword': keyword,
                    'title': title_text,
                    'link': real_link
                })

            safe_keyword = re.sub(r'[<>:"/\\|?*]', '_', keyword)
            with open(f"{safe_keyword}_results.json", "w", encoding="utf-8") as f:
                json.dump(search_results, f, ensure_ascii=False, indent=2)
            print(f"搜索结果已保存至 {safe_keyword}_results.json")

        print(f"\n开始检查 '{keyword}' 的搜索结果链接：")
        for i, result in enumerate(search_results, 1):
            link = result['link']
            print(f"\n{i}. 搜索关键词: {keyword}")
            print(f"   原URL: {link}")
            logging.info(f"检查链接: {link}")
            if link == '无链接':
                print(f"   内容类型: 未知")
                print(f"   页面标题: ")
                print(f"   关键词: ")
                print(f"   描述: ")
                print(f"   跳转后URL: {link}")
                print(f"   跳转后内容类型: 未知")
                print(f"   跳转后页面标题: ")
                print(f"   跳转后关键词: ")
                print(f"   跳转后描述: ")
                result['initial_metadata'] = {
                    'title': '',
                    'keywords': '',
                    'description': ''
                }
                result['initial_content_type'] = '未知'
                result['final_url'] = link
                result['metadata'] = {
                    'title': '',
                    'keywords': '',
                    'description': ''
                }
                result['final_content_type'] = '未知'
                continue

            max_retries = 3
            for attempt in range(max_retries):
                try:
                    driver_no_js.get(link)
                    WebDriverWait(driver_no_js, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    initial_html = driver_no_js.page_source
                    soup = BeautifulSoup(initial_html, 'html.parser')
                    initial_title = soup.find('title').text.strip() if soup.find('title') else ''
                    initial_keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
                    initial_keywords = initial_keywords_meta['content'].strip() if initial_keywords_meta else ''
                    initial_description_meta = soup.find('meta', attrs={'name': 'description'})
                    initial_description = initial_description_meta[
                        'content'].strip() if initial_description_meta else ''
                    initial_content_type = detect_content_type(link, initial_html)

                    print(f"   内容类型: {initial_content_type}")
                    print(f"   页面标题: {initial_title}")
                    print(f"   关键词: {initial_keywords}")
                    print(f"   描述: {initial_description}")

                    result['initial_metadata'] = {
                        'title': initial_title,
                        'keywords': initial_keywords,
                        'description': initial_description
                    }
                    result['initial_content_type'] = initial_content_type
                    break
                except TimeoutException:
                    print(f"   初始URL加载超时（5秒未响应，第 {attempt + 1}/{max_retries} 次尝试），刷新页面...")
                    logging.error(f"初始URL加载超时: {link} （第 {attempt + 1} 次尝试）")
                    if attempt == max_retries - 1:
                        print("   达到最大重试次数，跳过初始URL元数据提取...")
                        result['initial_metadata'] = {
                            'title': '',
                            'keywords': '',
                            'description': ''
                        }
                        result['initial_content_type'] = '未知'
                        break
                    driver_no_js.refresh()
                    time.sleep(random.uniform(1, 2))
                except Exception as e:
                    print(f"   初始URL访问失败: {e}")
                    logging.error(f"初始URL访问失败: {e}")
                    result['initial_metadata'] = {
                        'title': '',
                        'keywords': '',
                        'description': ''
                    }
                    result['initial_content_type'] = '未知'
                    break

            for attempt in range(max_retries):
                try:
                    driver_with_js.get(link)
                    WebDriverWait(driver_with_js, 5).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                    time.sleep(3)
                    WebDriverWait(driver_with_js, 10).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    final_url = driver_with_js.current_url
                    print(f"   跳转后URL: {final_url}")

                    if is_same_content_url(link, final_url):
                        print("   跳转后URL与初始URL实质相同，仅记录初始数据")
                        print(f"   跳转后内容类型: {result['initial_content_type']}")
                        print(f"   跳转后页面标题: {result['initial_metadata']['title']}")
                        print(f"   跳转后关键词: {result['initial_metadata']['keywords']}")
                        print(f"   跳转后描述: {result['initial_metadata']['description']}")
                        result['final_url'] = final_url
                        result['metadata'] = result['initial_metadata']
                        result['final_content_type'] = result['initial_content_type']
                    else:
                        final_html = driver_with_js.page_source
                        soup = BeautifulSoup(final_html, 'html.parser')
                        final_title = soup.find('title').text.strip() if soup.find('title') else ''
                        final_keywords_meta = soup.find('meta', attrs={'name': 'keywords'})
                        final_keywords = final_keywords_meta['content'].strip() if final_keywords_meta else ''
                        final_description_meta = soup.find('meta', attrs={'name': 'description'})
                        final_description = final_description_meta[
                            'content'].strip() if final_description_meta else ''
                        final_content_type = detect_content_type(final_url, final_html)

                        print(f"   跳转后内容类型: {final_content_type}")
                        print(f"   跳转后页面标题: {final_title}")
                        print(f"   跳转后关键词: {final_keywords}")
                        print(f"   跳转后描述: {final_description}")

                        result['final_url'] = final_url
                        result['metadata'] = {
                            'title': final_title,
                            'keywords': final_keywords,
                            'description': final_description
                        }
                        result['final_content_type'] = final_content_type
                    break
                except TimeoutException:
                    print(f"   跳转后URL加载超时（5秒未响应，第 {attempt + 1}/{max_retries} 次尝试），刷新页面...")
                    logging.error(f"跳转后URL加载超时: {link} （第 {attempt + 1} 次尝试）")
                    if attempt == max_retries - 1:
                        print("   达到最大重试次数，跳过此链接...")
                        result['final_url'] = link
                        result['metadata'] = {
                            'title': '',
                            'keywords': '',
                            'description': ''
                        }
                        result['final_content_type'] = '未知'
                        break
                    driver_with_js.refresh()
                    time.sleep(random.uniform(1, 2))
                except Exception as e:
                    print(f"   跳转后URL访问失败: {e}")
                    logging.error(f"跳转后URL访问失败: {e}")
                    result['final_url'] = link
                    result['metadata'] = {
                        'title': '',
                        'keywords': '',
                        'description': ''
                    }
                    result['final_content_type'] = '未知'
                    break

            # 保存原始 URL，Excel 会自动识别为超链接（用户手动编辑后按 Enter 也会触发）
            excel_row = {
                '搜索关键词': result['original_keyword'],
                '内容类型': result['initial_content_type'],
                '网址': result['link'],  # 直接保存原始 URL，Excel 会自动识别
                '页面标题': result['initial_metadata']['title'] if result['initial_metadata']['title'] else '无',
                '关键词': result['initial_metadata']['keywords'] if result['initial_metadata']['keywords'] else '无',
                '描述': result['initial_metadata']['description'] if result['initial_metadata']['description'] else '无',
                '跳转后的内容类型': result['final_content_type'] if not is_same_content_url(result['link'], result[
                    'final_url']) else '',
                '跳转后的网址': result['final_url'] if not is_same_content_url(result['link'],
                                                                               result['final_url']) else '',
                # 直接保存原始 URL，Excel 会自动识别
                '跳转后的页面标题': result['metadata']['title'] if (not is_same_content_url(result['link'], result['final_url']) and result['metadata']['title']) else '无' if not is_same_content_url(result['link'], result['final_url']) else '',
                '跳转后的关键词': result['metadata']['keywords'] if (not is_same_content_url(result['link'], result['final_url']) and result['metadata']['keywords']) else '无' if not is_same_content_url(result['link'], result['final_url']) else '',
                '跳转后的描述': result['metadata']['description'] if (not is_same_content_url(result['link'], result['final_url']) and result['metadata']['description']) else '无' if not is_same_content_url(result['link'], result['final_url']) else ''
            }
            all_excel_data.append(excel_row)

        # 保存原始 URL 到 JSON
        safe_keyword = re.sub(r'[<>:"/\\|?*]', '_', keyword)
        with open(f"{safe_keyword}_results_with_metadata.json", "w", encoding="utf-8") as f:
            json.dump(search_results, f, ensure_ascii=False, indent=2)
        print(f"包含元数据的搜索结果已保存至 {safe_keyword}_results_with_metadata.json")

        # 写入或追加到 Excel 文件
        df = pd.DataFrame(all_excel_data)
        if first_write:
            df.to_excel(excel_file, index=False, engine='openpyxl')
            first_write = False
        else:
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        print(f"关键词 '{keyword}' 的搜索结果已保存至 {excel_file}")

    # 加载 Excel 文件并应用格式
    wb = load_workbook(excel_file)
    ws = wb.active

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 设置“描述”列右对齐
    desc_col_idx = list(df.columns).index('描述') + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=desc_col_idx, max_col=desc_col_idx):
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    # 设置“跳转后的内容类型”列为黄色填充
    final_content_type_col_idx = list(df.columns).index('跳转后的内容类型') + 1
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=final_content_type_col_idx,
                            max_col=final_content_type_col_idx):
        for cell in row:
            cell.fill = yellow_fill

    # 保存格式化后的 Excel 文件
    wb.save(excel_file)
    print(f"所有搜索结果已保存至 {excel_file} 并完成格式化")

except WebDriverException as e:
    print(f"浏览器相关错误: {e}")
    logging.error(f"浏览器相关错误: {e}")
except Exception as e:
    print(f"出错: {e}")
    logging.error(f"出错: {e}")
finally:
    # 打印当前时间到分钟
    print(f"\n运行结束时间: {time.strftime('%Y-%m-%d %H:%M')}")

    if driver_no_js:
        print("\n关闭禁用 JavaScript 的浏览器...")
        logging.info("关闭禁用 JavaScript 的浏览器")
        time.sleep(2)
        driver_no_js.quit()
    else:
        print("\n禁用 JavaScript 的浏览器未成功初始化，跳过关闭。")
        logging.warning("禁用 JavaScript 的浏览器未成功初始化，跳过关闭")

    if driver_with_js:
        print("\n关闭启用 JavaScript 的浏览器...")
        logging.info("关闭启用 JavaScript 的浏览器")
        time.sleep(2)
        driver_with_js.quit()
    else:
        print("\n启用 JavaScript 的浏览器未成功初始化，跳过关闭。")
        logging.warning("启用 JavaScript 的浏览器未成功初始化，跳过关闭")